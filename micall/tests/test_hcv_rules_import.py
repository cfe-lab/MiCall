from argparse import Namespace
from copy import copy
from unittest import TestCase

from io import StringIO

from openpyxl import Workbook

from micall.utils.hcv_rules_import import (
    load_references,
    WorksheetReader,
    MonitoredPositionsReader,
    FoldRangesReader,
    RulesWriter,
    Range,
)

REFERENCES = load_references()


def create_worksheet(title, row_data):
    """Create a worksheet from a list of lists of cell values.

    Use these special cell values:
    '' - empty cell
    '<' - merge cell with left neighbour
    :param str title: the title for the worksheet
    :param list row_data: a list of lists of cell values
    """
    wb = Workbook()
    ws = wb.create_sheet(title)
    for i, row in enumerate(row_data, 1):
        row = [None if cell == "" else cell for cell in row]
        ws.append(row)
        start_column = None
        for j, cell in enumerate(row + [None], 1):
            if cell == "<":
                if start_column is None:
                    start_column = j - 1
            elif start_column is not None:
                end_column = j - 1
                ws.merge_cells(
                    start_row=i,
                    end_row=i,
                    start_column=start_column,
                    end_column=end_column,
                )
                start_column = None
    return ws


class RangeTest(TestCase):
    def test_lower_range_not_equal(self):
        changes = []
        invalid_fold_shifts = []
        r = Range(
            "< 2.5x FS, likely susceptible",
            "> 100x FS, resistance likely",
            "GTX Drugisvir",
            changes,
            invalid_fold_shifts,
        )
        expected_changes = [
            "GTX Drugisvir Y43M "
            "(2.5x/no in < 2.5, > 100 resistance possible) "
            "but is likely susceptible",
            "GTX Drugisvir P44I "
            "(2.6x/no in < 2.5, > 100 resistance possible) "
            "but is likely susceptible",
        ]
        expected_invalid_fold_shifts = []

        r.validate_phenotype("L42R", "likely susceptible", "2.4x")
        r.validate_phenotype("Y43M", "likely susceptible", "2.5x")
        r.validate_phenotype("P44I", "likely susceptible", "2.6x")

        self.assertEqual(expected_changes, changes)
        self.assertEqual(expected_invalid_fold_shifts, invalid_fold_shifts)

    def test_lower_range_equal(self):
        changes = []
        invalid_fold_shifts = []
        r = Range(
            "<= 2.5x FS, likely susceptible",
            "> 100x FS, resistance likely",
            "GTX Drugisvir",
            changes,
            invalid_fold_shifts,
        )
        expected_changes = [
            "GTX Drugisvir P44I "
            "(2.6x/no in <= 2.5, > 100 resistance possible) "
            "but is likely susceptible"
        ]
        expected_invalid_fold_shifts = []

        r.validate_phenotype("L42R", "likely susceptible", "2.4x")
        r.validate_phenotype("Y43M", "likely susceptible", "2.5x")
        r.validate_phenotype("P44I", "likely susceptible", "2.6x")

        self.assertEqual(expected_changes, changes)
        self.assertEqual(expected_invalid_fold_shifts, invalid_fold_shifts)

    def test_upper_range_not_equal(self):
        changes = []
        invalid_fold_shifts = []
        r = Range(
            "< 2.5x FS, likely susceptible",
            "> 100x FS, resistance likely",
            "GTX Drugisvir",
            changes,
            invalid_fold_shifts,
        )
        expected_changes = [
            "GTX Drugisvir L42R "
            "(99.9x/no in < 2.5, > 100 resistance possible) "
            "but is resistance likely",
            "GTX Drugisvir Y43M "
            "(100.0x/no in < 2.5, > 100 resistance possible) "
            "but is resistance likely",
        ]
        expected_invalid_fold_shifts = []

        r.validate_phenotype("L42R", "resistance likely", "99.9x")
        r.validate_phenotype("Y43M", "resistance likely", "100.0x")
        r.validate_phenotype("P44I", "resistance likely", "100.1x")

        self.assertEqual(expected_changes, changes)
        self.assertEqual(expected_invalid_fold_shifts, invalid_fold_shifts)

    def test_upper_range_equal(self):
        changes = []
        invalid_fold_shifts = []
        r = Range(
            "< 2.5x FS, likely susceptible",
            ">= 100x FS, resistance likely",
            "GTX Drugisvir",
            changes,
            invalid_fold_shifts,
        )
        expected_changes = [
            "GTX Drugisvir L42R "
            "(99.9x/no in < 2.5, >= 100 resistance possible) "
            "but is resistance likely"
        ]
        expected_invalid_fold_shifts = []

        r.validate_phenotype("L42R", "resistance likely", "99.9x")
        r.validate_phenotype("Y43M", "resistance likely", "100.0x")
        r.validate_phenotype("P44I", "resistance likely", "100.1x")

        self.assertEqual(expected_changes, changes)
        self.assertEqual(expected_invalid_fold_shifts, invalid_fold_shifts)

    def test_fold_shift_comparison(self):
        changes = []
        invalid_fold_shifts = []
        r = Range(
            "< 2.5x FS, likely susceptible",
            "> 100x FS, resistance likely",
            "GTX Drugisvir",
            changes,
            invalid_fold_shifts,
        )
        expected_changes = []
        expected_invalid_fold_shifts = []

        r.validate_phenotype("Y43M", "resistance likely", ">100.0x")
        r.validate_phenotype("P44I", "likely susceptible", "<2.5x")

        self.assertEqual(expected_changes, changes)
        self.assertEqual(expected_invalid_fold_shifts, invalid_fold_shifts)

    def test_ignore_commas(self):
        changes = []
        invalid_fold_shifts = []
        r = Range(
            "< 2.5x FS, likely susceptible",
            "> 100x FS, resistance likely",
            "GTX Drugisvir",
            changes,
            invalid_fold_shifts,
        )
        expected_changes = []
        expected_invalid_fold_shifts = []

        r.validate_phenotype("Y43M", "resistance likely", "1,000.0x")

        self.assertEqual(expected_changes, changes)
        self.assertEqual(expected_invalid_fold_shifts, invalid_fold_shifts)

    def test_lower_limit_invalid(self):
        changes = []
        invalid_fold_shifts = []
        r = Range(
            "< 2.5 times FS, likely susceptible",
            ">= 100x FS, resistance likely",
            "GTX Drugisvir",
            changes,
            invalid_fold_shifts,
        )
        expected_changes = []
        expected_invalid_fold_shifts = [
            "GTX Drugisvir lower fold shift: < 2.5 times FS, likely susceptible"
        ]

        r.validate_phenotype("L42R", "resistance likely", "99.9x")
        r.validate_phenotype("Y43M", "resistance likely", "100.0x")
        r.validate_phenotype("P44I", "resistance likely", "100.1x")

        self.assertEqual(expected_changes, changes)
        self.assertEqual(expected_invalid_fold_shifts, invalid_fold_shifts)

    def test_upper_limit_invalid(self):
        changes = []
        invalid_fold_shifts = []
        r = Range(
            "< 2.5 times FS, likely susceptible",
            ">= 100x FS, resistance spooky",
            "GTX Drugisvir",
            changes,
            invalid_fold_shifts,
        )
        expected_changes = []
        expected_invalid_fold_shifts = [
            "GTX Drugisvir lower fold shift: < 2.5 times FS, likely susceptible",
            "GTX Drugisvir upper fold shift: >= 100x FS, resistance spooky",
        ]

        r.validate_phenotype("L42R", "resistance likely", "99.9x")
        r.validate_phenotype("Y43M", "resistance likely", "100.0x")
        r.validate_phenotype("P44I", "resistance likely", "100.1x")

        self.assertEqual(expected_changes, changes)
        self.assertEqual(expected_invalid_fold_shifts, invalid_fold_shifts)

    def test_upper_comparison_invalid(self):
        changes = []
        invalid_fold_shifts = []
        r = Range(
            "< 2.5x FS, likely susceptible",
            "= 100x FS, resistance likely",
            "GTX Drugisvir",
            changes,
            invalid_fold_shifts,
        )
        expected_changes = []
        expected_invalid_fold_shifts = [
            "GTX Drugisvir upper fold shift: = 100x FS, resistance likely"
        ]

        r.validate_phenotype("L42R", "resistance likely", "99.9x")
        r.validate_phenotype("Y43M", "resistance likely", "100.0x")
        r.validate_phenotype("P44I", "resistance likely", "100.1x")

        self.assertEqual(expected_changes, changes)
        self.assertEqual(expected_invalid_fold_shifts, invalid_fold_shifts)

    def test_lower_comparison_invalid(self):
        changes = []
        invalid_fold_shifts = []
        r = Range(
            "= 2.5x FS, likely susceptible",
            ">= 100x FS, resistance likely",
            "GTX Drugisvir",
            changes,
            invalid_fold_shifts,
        )
        expected_changes = []
        expected_invalid_fold_shifts = [
            "GTX Drugisvir lower fold shift: = 2.5x FS, likely susceptible"
        ]

        r.validate_phenotype("L42R", "resistance likely", "99.9x")
        r.validate_phenotype("Y43M", "resistance likely", "100.0x")
        r.validate_phenotype("P44I", "resistance likely", "100.1x")

        self.assertEqual(expected_changes, changes)
        self.assertEqual(expected_invalid_fold_shifts, invalid_fold_shifts)

    def test_invalid_fold_shift(self):
        changes = []
        invalid_fold_shifts = []
        r = Range(
            "< 2.5x FS, very strange",
            ">= 100x FS, resistance likely",
            "GTX Drugisvir",
            changes,
            invalid_fold_shifts,
        )
        expected_changes = []
        expected_invalid_fold_shifts = [
            "GTX Drugisvir lower fold shift: < 2.5x FS, very strange",
            "GTX Drugisvir L42R 'twenty'",
        ]

        r.validate_phenotype("L42R", "resistance likely", "twenty")

        self.assertEqual(expected_changes, changes)
        self.assertEqual(expected_invalid_fold_shifts, invalid_fold_shifts)

    def test_susceptible_range(self):
        changes = []
        invalid_fold_shifts = []
        r = Range(
            "< 20x FS, likely susceptible",
            ">= 100x FS, resistance likely",
            "GTX Drugisvir",
            changes,
            invalid_fold_shifts,
        )
        expected_changes = [
            "GTX Drugisvir Y43M "
            "(10-15x/no in < 20, >= 100 likely susceptible) "
            "but is resistance possible"
        ]
        expected_invalid_fold_shifts = []

        r.validate_phenotype("L42R", "likely susceptible", "1-20x")
        r.validate_phenotype("Y43M", "resistance possible", "10-15x")

        self.assertEqual(expected_changes, changes)
        self.assertEqual(expected_invalid_fold_shifts, invalid_fold_shifts)

    def test_resistant_range(self):
        changes = []
        invalid_fold_shifts = []
        r = Range(
            "< 20x FS, likely susceptible",
            ">= 100x FS, resistance likely",
            "GTX Drugisvir",
            changes,
            invalid_fold_shifts,
        )
        expected_changes = [
            "GTX Drugisvir Y43M "
            "(100-9000x/no in < 20, >= 100 resistance likely) "
            "but is resistance possible"
        ]
        expected_invalid_fold_shifts = []

        r.validate_phenotype("L42R", "resistance likely", "100-2000x")
        r.validate_phenotype("Y43M", "resistance possible", "100-9000x")

        self.assertEqual(expected_changes, changes)
        self.assertEqual(expected_invalid_fold_shifts, invalid_fold_shifts)

    def test_possible_range(self):
        changes = []
        invalid_fold_shifts = []
        r = Range(
            "< 20x FS, likely susceptible",
            ">= 100x FS, resistance likely",
            "GTX Drugisvir",
            changes,
            invalid_fold_shifts,
        )
        expected_changes = [
            "GTX Drugisvir L42R "
            "(20-99.9x/no in < 20, >= 100 resistance possible) "
            "but is resistance likely"
        ]
        expected_invalid_fold_shifts = []

        r.validate_phenotype("L42R", "resistance likely", "20-99.9x")
        r.validate_phenotype("Y43M", "resistance possible", "50-90x")

        self.assertEqual(expected_changes, changes)
        self.assertEqual(expected_invalid_fold_shifts, invalid_fold_shifts)

    def test_possible_range_overlap(self):
        changes = []
        invalid_fold_shifts = []
        r = Range(
            "1-20x FS, likely susceptible",
            "100-1000x FS, resistance likely",
            "GTX Drugisvir",
            changes,
            invalid_fold_shifts,
        )
        expected_changes = []
        expected_invalid_fold_shifts = []

        r.validate_phenotype("L42R", "likely susceptible", "10-20x")
        r.validate_phenotype("Y43M", "resistance possible", "20-100x")
        r.validate_phenotype("M44Y", "resistance likely", "100-200x")

        self.assertEqual(expected_changes, changes)
        self.assertEqual(expected_invalid_fold_shifts, invalid_fold_shifts)

    def test_ambiguous_range(self):
        changes = []
        invalid_fold_shifts = []
        r = Range(
            "< 20x FS, likely susceptible",
            ">= 100x FS, resistance likely",
            "GTX Drugisvir",
            changes,
            invalid_fold_shifts,
        )
        expected_changes = [
            "GTX Drugisvir L42R "
            "(20-120x/no in < 20, >= 100 ambiguous) "
            "but is resistance likely",
            "GTX Drugisvir Y43M "
            "(5-90x/no in < 20, >= 100 ambiguous) "
            "but is resistance possible",
        ]
        expected_invalid_fold_shifts = []

        r.validate_phenotype("L42R", "resistance likely", "20-120x")
        r.validate_phenotype("Y43M", "resistance possible", "5-90x")

        self.assertEqual(expected_changes, changes)
        self.assertEqual(expected_invalid_fold_shifts, invalid_fold_shifts)

    def test_clinical_increases_resistance(self):
        changes = []
        invalid_fold_shifts = []
        r = Range(
            "< 20x FS, likely susceptible",
            ">= 100x FS, resistance likely",
            "GTX Drugisvir",
            changes,
            invalid_fold_shifts,
        )
        expected_changes = []
        expected_invalid_fold_shifts = []

        r.validate_phenotype("L42R", "resistance possible", "10x", clinical_ras="yes")

        self.assertEqual(expected_changes, changes)
        self.assertEqual(expected_invalid_fold_shifts, invalid_fold_shifts)

    def test_clinical_at_max(self):
        changes = []
        invalid_fold_shifts = []
        r = Range(
            "< 20x FS, likely susceptible",
            ">= 100x FS, resistance likely",
            "GTX Drugisvir",
            changes,
            invalid_fold_shifts,
        )
        expected_changes = []
        expected_invalid_fold_shifts = []

        r.validate_phenotype("L42R", "resistance likely", "100x", clinical_ras="yes")

        self.assertEqual(expected_changes, changes)
        self.assertEqual(expected_invalid_fold_shifts, invalid_fold_shifts)

    def test_minimal_constructor(self):
        r = Range("<20 FS, likely susceptible", ">100 FS, resistance likely")
        expected_changes = [
            "L42R (200x/no in <20, >100 resistance likely) but is resistance possible"
        ]
        expected_invalid_fold_shifts = []

        r.validate_phenotype("L42R", "resistance possible", "200x")

        self.assertEqual(expected_changes, r.changes)
        self.assertEqual(expected_invalid_fold_shifts, r.invalid_fold_shifts)

    def test_repr_and_str(self):
        r = Range(
            "<20 FS, likely susceptible", ">100 FS, resistance likely", "GTX Drugisvir"
        )

        self.assertEqual(
            "Range('<20 FS, likely susceptible', '>100 FS, resistance likely')", repr(r)
        )
        self.assertEqual("<20, >100", str(r))

    def test_equality(self):
        r1 = Range("<2 FS, likely susceptible", ">8 FS, resistance likely")
        r2 = Range("< 2 FS, likely susceptible", ">8.0x FS, resistance likely")
        r3 = Range("<1 FS, likely susceptible", ">8 FS, resistance likely")
        r4 = Range("<2 FS, likely susceptible", ">9 FS, resistance likely")
        r5 = Range("<=2 FS, likely susceptible", ">8 FS, resistance likely")
        r6 = Range("<2 FS, likely susceptible", ">=8 FS, resistance likely")
        r7 = Range(
            "<2 FS, likely susceptible", ">8 FS, resistance likely", "GTX Drugisvir"
        )
        r8 = Range("bogus", "range")

        self.assertEqual(r1, r2)
        self.assertNotEqual(r1, r3)
        self.assertNotEqual(r1, r4)
        self.assertNotEqual(r1, r5)
        self.assertNotEqual(r1, r6)
        self.assertEqual(r1, r7)
        self.assertEqual(r8, r8)


class WorksheetReaderTest(TestCase):
    def setUp(self):
        self.expected_errors = ""

    def assertReads(self, expected_entries, worksheets, *footer_readers):
        errors = StringIO()
        reader = WorksheetReader(worksheets, *footer_readers)

        entries = list(reader)
        reader.write_errors(errors)

        self.assertEqual(self.expected_errors, errors.getvalue())
        self.assertEqual(expected_entries, entries)

    def test_one_sheet(self):
        row_data = [
            ["", "Ignored header row"],
            ["", "Example1 drug", "<", "<", "Example2", "<", "<"],
            ["", "a", "Phenotype", "b", "a", "b", "Phenotype"],
            ["WT", "", "likely susceptible", "", "", "", "likely susceptible"],
            ["V36A", "", "resistance likely", "", "", "", "likely susceptible"],
            ["T40A", "", "likely susceptible", "", "", "", "resistance possible"],
        ]
        worksheets = [create_worksheet("NS3_GT1a", row_data)]
        expected_section1 = Namespace(drug_name="Example1", sheet_name="NS3_GT1a")
        expected_section2 = Namespace(drug_name="Example2", sheet_name="NS3_GT1a")
        expected_entries = [
            Namespace(
                mutation="WT",
                section=expected_section1,
                phenotype="likely susceptible",
                a=None,
                b=None,
            ),
            Namespace(
                mutation="V36A",
                section=expected_section1,
                phenotype="resistance likely",
                a=None,
                b=None,
            ),
            Namespace(
                mutation="T40A",
                section=expected_section1,
                phenotype="likely susceptible",
                a=None,
                b=None,
            ),
            Namespace(
                mutation="WT",
                section=expected_section2,
                phenotype="likely susceptible",
                a=None,
                b=None,
            ),
            Namespace(
                mutation="V36A",
                section=expected_section2,
                phenotype="likely susceptible",
                a=None,
                b=None,
            ),
            Namespace(
                mutation="T40A",
                section=expected_section2,
                phenotype="resistance possible",
                a=None,
                b=None,
            ),
        ]

        self.assertReads(expected_entries, worksheets)

    def test_two_sheets(self):
        row_data1 = [
            ["", "Ignored header row"],
            ["", "Example1 drug", "<", "<"],
            ["", "a", "Phenotype", "b"],
            ["WT", "", "likely susceptible", ""],
            ["V36A", "", "resistance likely", ""],
            ["T40A", "", "likely susceptible", ""],
        ]
        row_data2 = [
            ["", "Example2", "<", "<"],
            ["", "a", "b", "Phenotype"],
            ["WT", "", "", "likely susceptible"],
            ["V36A", "", "", "likely susceptible"],
            ["T40A", "", "", "resistance possible"],
        ]
        worksheets = [
            create_worksheet("NS3_GT1a", row_data1),
            create_worksheet("NS3_GT1b", row_data2),
        ]
        expected_section1 = Namespace(drug_name="Example1", sheet_name="NS3_GT1a")
        expected_section2 = Namespace(drug_name="Example2", sheet_name="NS3_GT1b")
        expected_entries = [
            Namespace(
                mutation="WT",
                section=expected_section1,
                phenotype="likely susceptible",
                a=None,
                b=None,
            ),
            Namespace(
                mutation="V36A",
                section=expected_section1,
                phenotype="resistance likely",
                a=None,
                b=None,
            ),
            Namespace(
                mutation="T40A",
                section=expected_section1,
                phenotype="likely susceptible",
                a=None,
                b=None,
            ),
            Namespace(
                mutation="WT",
                section=expected_section2,
                phenotype="likely susceptible",
                a=None,
                b=None,
            ),
            Namespace(
                mutation="V36A",
                section=expected_section2,
                phenotype="likely susceptible",
                a=None,
                b=None,
            ),
            Namespace(
                mutation="T40A",
                section=expected_section2,
                phenotype="resistance possible",
                a=None,
                b=None,
            ),
        ]

        self.assertReads(expected_entries, worksheets)

    def test_missing_wild_type(self):
        row_data1 = [
            ["", "Ignored header row"],
            ["", "Example1 drug", "<", "<"],
            ["", "a", "Phenotype", "b"],
            ["V36A", "", "resistance likely", ""],
            ["T40A", "", "likely susceptible", ""],
        ]
        row_data2 = [
            ["", "Example2", "<", "<"],
            ["", "a", "b", "Phenotype"],
            ["WT", "", "", "likely susceptible"],
            ["V36A", "", "", "likely susceptible"],
            ["T40A", "", "", "resistance possible"],
        ]
        worksheets = [
            create_worksheet("NS3_GT1a", row_data1),
            create_worksheet("NS3_GT1b", row_data2),
        ]
        expected_section2 = Namespace(drug_name="Example2", sheet_name="NS3_GT1b")
        expected_entries = [
            Namespace(
                mutation="WT",
                section=expected_section2,
                phenotype="likely susceptible",
                a=None,
                b=None,
            ),
            Namespace(
                mutation="V36A",
                section=expected_section2,
                phenotype="likely susceptible",
                a=None,
                b=None,
            ),
            Namespace(
                mutation="T40A",
                section=expected_section2,
                phenotype="resistance possible",
                a=None,
                b=None,
            ),
        ]
        self.expected_errors = """\
No wild type found in NS3_GT1a.
"""

        self.assertReads(expected_entries, worksheets)

    def test_blank(self):
        row_data = [
            ["", "Example", "<"],
            ["", "Phenotype"],
            ["WT", "likely susceptible"],
            ["V36A", ""],
            ["T40A", "resistance possible"],
        ]
        worksheets = [create_worksheet("NS3_GT1a", row_data)]
        expected_section = Namespace(drug_name="Example", sheet_name="NS3_GT1a")
        expected_entries = [
            Namespace(
                mutation="WT", section=expected_section, phenotype="likely susceptible"
            ),
            Namespace(
                mutation="T40A",
                section=expected_section,
                phenotype="resistance possible",
            ),
        ]

        self.assertReads(expected_entries, worksheets)

    def test_mutations_stripped(self):
        row_data = [
            ["", "Example", "<"],
            ["", "Phenotype"],
            ["WT", "likely susceptible"],
            [" V36A ", "likely susceptible"],
        ]
        worksheets = [create_worksheet("NS3_GT1a", row_data)]
        expected_section = Namespace(drug_name="Example", sheet_name="NS3_GT1a")
        expected_entries = [
            Namespace(
                mutation="WT", section=expected_section, phenotype="likely susceptible"
            ),
            Namespace(
                mutation="V36A",
                section=expected_section,
                phenotype="likely susceptible",
            ),
        ]
        self.assertReads(expected_entries, worksheets)

    def test_mutation_deletion(self):
        row_data = [
            ["", "Example", "<"],
            ["", "Phenotype"],
            ["WT", "likely susceptible"],
            ["V36del", "likely susceptible"],
        ]
        worksheets = [create_worksheet("NS3_GT1a", row_data)]
        expected_section = Namespace(drug_name="Example", sheet_name="NS3_GT1a")
        expected_entries = [
            Namespace(
                mutation="WT", section=expected_section, phenotype="likely susceptible"
            ),
            Namespace(
                mutation="V36d",
                section=expected_section,
                phenotype="likely susceptible",
            ),
        ]
        self.assertReads(expected_entries, worksheets)

    def test_other_merged_section(self):
        row_data = [
            ["", "Example", "<", "<"],
            ["", "a", "b", "Phenotype"],
            ["WT", "", "", "likely susceptible"],
            ["V36A", "", "", "likely susceptible"],
            ["T40A", "", "", "resistance possible"],
            ["", "Other merged section", "<"],
        ]
        worksheets = [create_worksheet("NS3_GT1b", row_data)]
        expected_section = Namespace(drug_name="Example", sheet_name="NS3_GT1b")
        expected_entries = [
            Namespace(
                mutation="WT",
                section=expected_section,
                phenotype="likely susceptible",
                a=None,
                b=None,
            ),
            Namespace(
                mutation="V36A",
                section=expected_section,
                phenotype="likely susceptible",
                a=None,
                b=None,
            ),
            Namespace(
                mutation="T40A",
                section=expected_section,
                phenotype="resistance possible",
                a=None,
                b=None,
            ),
        ]

        self.assertReads(expected_entries, worksheets)

    def test_monitored_positions(self):
        row_data = [
            ["", "Example", "<"],
            ["", "Phenotype"],
            ["WT", "likely susceptible"],
            ["V36A", "likely susceptible"],
            ["T40A", "resistance possible"],
            ["", "Positions monitored:"],
            ["", "36, 99"],
        ]
        worksheets = [create_worksheet("NS3_GT1b", row_data)]
        expected_section = Namespace(
            drug_name="Example", sheet_name="NS3_GT1b", monitored_positions=[36, 99]
        )
        expected_entries = [
            Namespace(
                mutation="WT", section=expected_section, phenotype="likely susceptible"
            ),
            Namespace(
                mutation="V36A",
                section=expected_section,
                phenotype="likely susceptible",
            ),
            Namespace(
                mutation="T40A",
                section=expected_section,
                phenotype="resistance possible",
            ),
        ]

        self.assertReads(expected_entries, worksheets, MonitoredPositionsReader())

    def test_monitored_position_int(self):
        row_data = [
            ["", "Example", "<"],
            ["", "Phenotype"],
            ["WT", "likely susceptible"],
            ["V36A", "likely susceptible"],
            ["T40A", "resistance possible"],
            ["", "Positions monitored:"],
            ["", 36],
        ]
        worksheets = [create_worksheet("NS3_GT1b", row_data)]
        expected_section = Namespace(
            drug_name="Example", sheet_name="NS3_GT1b", monitored_positions=[36]
        )
        expected_entries = [
            Namespace(
                mutation="WT", section=expected_section, phenotype="likely susceptible"
            ),
            Namespace(
                mutation="V36A",
                section=expected_section,
                phenotype="likely susceptible",
            ),
            Namespace(
                mutation="T40A",
                section=expected_section,
                phenotype="resistance possible",
            ),
        ]

        self.assertReads(expected_entries, worksheets, MonitoredPositionsReader())

    def test_no_monitored_positions(self):
        row_data = [
            ["", "Example", "<"],
            ["", "Phenotype"],
            ["WT", "likely susceptible"],
            ["V36A", "likely susceptible"],
            ["T40A", "resistance possible"],
            ["", "Nothing monitored."],
        ]
        worksheets = [create_worksheet("NS3_GT1b", row_data)]
        self.expected_errors = """\
No monitored positions for Example in NS3_GT1b.
"""
        expected_section = Namespace(
            drug_name="Example", sheet_name="NS3_GT1b", monitored_positions=[]
        )
        expected_entries = [
            Namespace(
                mutation="WT", section=expected_section, phenotype="likely susceptible"
            ),
            Namespace(
                mutation="V36A",
                section=expected_section,
                phenotype="likely susceptible",
            ),
            Namespace(
                mutation="T40A",
                section=expected_section,
                phenotype="resistance possible",
            ),
        ]

        self.assertReads(expected_entries, worksheets, MonitoredPositionsReader())

    def test_not_indicated(self):
        row_data = [
            ["", "Example", "<"],
            ["", "Phenotype"],
            ["WT", "likely susceptible"],
            ["V36A", "likely susceptible"],
            ["T40A", "resistance possible"],
            ["", "Nothing monitored."],
            ["", "Not indicated in Canada"],
        ]
        worksheets = [create_worksheet("NS3_GT1b", row_data)]
        expected_section = Namespace(
            drug_name="Example", sheet_name="NS3_GT1b", not_indicated=True
        )
        expected_entries = [Namespace(section=expected_section)]

        self.assertReads(expected_entries, worksheets)

    def test_obsolete_drug(self):
        row_data = [
            ["", "Boceprevir", "<"],
            ["", "Phenotype"],
            ["WT", "likely susceptible"],
            ["V36A", "likely susceptible"],
            ["T40A", "resistance possible"],
            ["", "Nothing monitored."],
            ["", "Not indicated in Canada"],
        ]
        worksheets = [create_worksheet("NS3_GT1b", row_data)]
        expected_entries = []

        self.assertReads(expected_entries, worksheets)

    def test_obsolete_drug_still_indicated(self):
        row_data = [
            ["", "Boceprevir", "<"],
            ["", "Phenotype"],
            ["WT", "likely susceptible"],
            ["V36A", "likely susceptible"],
            ["T40A", "resistance possible"],
            ["", "Nothing monitored."],
        ]  # No entry saying "Not indicated".
        worksheets = [create_worksheet("NS3_GT1b", row_data)]
        expected_section = Namespace(drug_name="Boceprevir", sheet_name="NS3_GT1b")
        expected_entries = [
            Namespace(
                mutation="WT", section=expected_section, phenotype="likely susceptible"
            ),
            Namespace(
                mutation="V36A",
                section=expected_section,
                phenotype="likely susceptible",
            ),
            Namespace(
                mutation="T40A",
                section=expected_section,
                phenotype="resistance possible",
            ),
        ]
        self.expected_errors = """\
Obsolete drugs still indicated: Boceprevir in NS3_GT1b.
"""

        self.assertReads(expected_entries, worksheets)

    def test_fold_shift_levels(self):
        row_data = [
            ["", "Example", "<", "<"],
            ["", "Phenotype"],
            ["WT", "likely susceptible"],
            ["V36A", "likely susceptible"],
            ["T40A", "resistance possible"],
            ["", "In vitro drug susceptibility:"],
            ["", "<20x FS, likely susceptible"],
            ["", "20-100x FS, resistance possible"],
            ["", ">100x FS, resistance likely"],
            ["", "", "Positions monitored:"],
            ["", "", "36, 99"],
        ]
        worksheets = [create_worksheet("NS3_GT1b", row_data)]
        expected_range = Range(
            "< 20x FS, likely susceptible", "> 100x FS, resistance likely"
        )
        expected_section = Namespace(
            drug_name="Example",
            sheet_name="NS3_GT1b",
            monitored_positions=[36, 99],
            fold_range=expected_range,
            range_range=expected_range,
        )
        expected_entries = [
            Namespace(
                mutation="WT", section=expected_section, phenotype="likely susceptible"
            ),
            Namespace(
                mutation="V36A",
                section=expected_section,
                phenotype="likely susceptible",
            ),
            Namespace(
                mutation="T40A",
                section=expected_section,
                phenotype="resistance possible",
            ),
        ]

        self.assertReads(
            expected_entries, worksheets, MonitoredPositionsReader(), FoldRangesReader()
        )

    def test_fold_shift_typos(self):
        row_data = [
            ["", "Example", "<", "<"],
            ["", "Phenotype"],
            ["WT", "likely susceptible"],
            ["V36A", "likely susceptible"],
            ["T40A", "resistance possible"],
            ["", "In virto Drug Susecptibility:"],
            ["", "<20x FS, likely susceptible"],
            ["", "20-100x FS, resistance possible"],
            ["", ">100x FS, resistance likely"],
        ]
        worksheets = [create_worksheet("NS3_GT1b", row_data)]
        expected_range = Range(
            "< 20x FS, likely susceptible", "> 100x FS, resistance likely"
        )
        expected_section = Namespace(
            drug_name="Example",
            sheet_name="NS3_GT1b",
            fold_range=expected_range,
            range_range=expected_range,
        )
        expected_entries = [
            Namespace(
                mutation="WT", section=expected_section, phenotype="likely susceptible"
            ),
            Namespace(
                mutation="V36A",
                section=expected_section,
                phenotype="likely susceptible",
            ),
            Namespace(
                mutation="T40A",
                section=expected_section,
                phenotype="resistance possible",
            ),
        ]

        self.assertReads(expected_entries, worksheets, FoldRangesReader())

    def test_fold_shift_absolute(self):
        row_data = [
            ["", "Example", "<", "<"],
            ["", "Phenotype"],
            ["WT", "likely susceptible"],
            ["V36A", "likely susceptible"],
            ["T40A", "resistance possible"],
            ["", "In vitro Drug Susceptibility:"],
            ["", "Absolute FS values -"],
            ["", "<20x FS, likely susceptible"],
            ["", "20-100x FS, resistance possible"],
            ["", ">100x FS, resistance likely"],
        ]
        worksheets = [create_worksheet("NS3_GT1b", row_data)]
        expected_range = Range(
            "< 20x FS, likely susceptible", "> 100x FS, resistance likely"
        )
        expected_section = Namespace(
            drug_name="Example",
            sheet_name="NS3_GT1b",
            fold_range=expected_range,
            range_range=expected_range,
        )
        expected_entries = [
            Namespace(
                mutation="WT", section=expected_section, phenotype="likely susceptible"
            ),
            Namespace(
                mutation="V36A",
                section=expected_section,
                phenotype="likely susceptible",
            ),
            Namespace(
                mutation="T40A",
                section=expected_section,
                phenotype="resistance possible",
            ),
        ]

        self.assertReads(expected_entries, worksheets, FoldRangesReader())

    def test_fold_shift_ranges(self):
        row_data = [
            ["", "Example", "<", "<"],
            ["", "Phenotype"],
            ["WT", "likely susceptible"],
            ["V36A", "likely susceptible"],
            ["T40A", "resistance possible"],
            ["", "In vitro Drug Susceptibility:"],
            ["", "Absolute FS values -"],
            ["", "<20x FS, likely susceptible"],
            ["", "20-100x FS, resistance possible"],
            ["", ">100x FS, resistance likely"],
            ["", "Range FS values -"],
            ["", "2.5-10x FS range, likely susceptible"],
            ["", "10-100x FS range, resistance possible"],
            ["", "100-1000x FS range, resistance likely"],
        ]
        worksheets = [create_worksheet("NS3_GT1b", row_data)]
        expected_fold_range = Range(
            "< 20x FS, likely susceptible", "> 100x FS, resistance likely"
        )
        expected_range_range = Range(
            "<= 10x FS, likely susceptible", ">= 100x FS, resistance likely"
        )
        expected_section = Namespace(
            drug_name="Example",
            sheet_name="NS3_GT1b",
            fold_range=expected_fold_range,
            range_range=expected_range_range,
        )
        expected_entries = [
            Namespace(
                mutation="WT", section=expected_section, phenotype="likely susceptible"
            ),
            Namespace(
                mutation="V36A",
                section=expected_section,
                phenotype="likely susceptible",
            ),
            Namespace(
                mutation="T40A",
                section=expected_section,
                phenotype="resistance possible",
            ),
        ]

        self.assertReads(expected_entries, worksheets, FoldRangesReader())

    def test_fold_shift_range_limits(self):
        row_data = [
            ["", "Example", "<", "<"],
            ["", "Phenotype"],
            ["WT", "likely susceptible"],
            ["V36A", "likely susceptible"],
            ["T40A", "resistance possible"],
            ["", "In vitro Drug Susceptibility:"],
            ["", "Absolute FS values -"],
            ["", "<20x FS, likely susceptible"],
            ["", "20-100x FS, resistance possible"],
            ["", ">100x FS, resistance likely"],
            ["", "Range FS values -"],
            ["", "<10x FS range, likely susceptible"],
            ["", "10-100x FS range, resistance possible"],
            ["", ">100x FS range, resistance likely"],
        ]
        worksheets = [create_worksheet("NS3_GT1b", row_data)]
        expected_fold_range = Range(
            "< 20x FS, likely susceptible", "> 100x FS, resistance likely"
        )
        expected_range_range = Range(
            "< 10x FS, likely susceptible", "> 100x FS, resistance likely"
        )
        expected_section = Namespace(
            drug_name="Example",
            sheet_name="NS3_GT1b",
            fold_range=expected_fold_range,
            range_range=expected_range_range,
        )
        expected_entries = [
            Namespace(
                mutation="WT", section=expected_section, phenotype="likely susceptible"
            ),
            Namespace(
                mutation="V36A",
                section=expected_section,
                phenotype="likely susceptible",
            ),
            Namespace(
                mutation="T40A",
                section=expected_section,
                phenotype="resistance possible",
            ),
        ]

        self.assertReads(expected_entries, worksheets, FoldRangesReader())

    def test_invalid_fold_shift(self):
        row_data = [
            ["", "Example", "<", "<"],
            ["", "Phenotype"],
            ["WT", "likely susceptible"],
            ["V36A", "likely susceptible"],
            ["T40A", "resistance possible"],
            ["", "In vitro drug susceptibility:"],
            ["", "less than 20x FS, moist"],
            ["", "20-100x FS, resistance possible"],
            ["", ">100x FS, resistance likely"],
        ]
        worksheets = [create_worksheet("NS3_GT1b", row_data)]
        self.expected_errors = """\
Example in NS3_GT1b lower fold shift: less than 20x FS, moist
"""
        expected_range = Range("less than 20x FS, moist", ">100x FS, resistance likely")
        expected_section = Namespace(
            drug_name="Example",
            sheet_name="NS3_GT1b",
            fold_range=expected_range,
            range_range=expected_range,
        )
        expected_entries = [
            Namespace(
                mutation="WT", section=expected_section, phenotype="likely susceptible"
            ),
            Namespace(
                mutation="V36A",
                section=expected_section,
                phenotype="likely susceptible",
            ),
            Namespace(
                mutation="T40A",
                section=expected_section,
                phenotype="resistance possible",
            ),
        ]

        self.assertReads(expected_entries, worksheets, FoldRangesReader())

    def test_heading_underscores(self):
        row_data = [
            ["", "Example", "<"],
            ["", "Fold-Shift", "Phenotype"],
            ["WT", "1x", "likely susceptible"],
            ["V36A", "2x", "likely susceptible"],
            ["T40A", "50x", "resistance possible"],
        ]
        worksheets = [create_worksheet("NS3_GT1b", row_data)]
        expected_section = Namespace(drug_name="Example", sheet_name="NS3_GT1b")
        expected_entries = [
            Namespace(
                mutation="WT",
                section=expected_section,
                phenotype="likely susceptible",
                fold_shift="1x",
            ),
            Namespace(
                mutation="V36A",
                section=expected_section,
                phenotype="likely susceptible",
                fold_shift="2x",
            ),
            Namespace(
                mutation="T40A",
                section=expected_section,
                phenotype="resistance possible",
                fold_shift="50x",
            ),
        ]

        self.assertReads(expected_entries, worksheets)

    def test_heading_stripped(self):
        row_data = [
            ["", "Example", "<"],
            ["", "Phenotype", "Comment*"],
            ["WT", "likely susceptible"],
            ["V36A", "likely susceptible"],
            ["T40A", "resistance possible", "unreliable"],
        ]
        worksheets = [create_worksheet("NS3_GT1b", row_data)]
        expected_section = Namespace(drug_name="Example", sheet_name="NS3_GT1b")
        expected_entries = [
            Namespace(
                mutation="WT",
                section=expected_section,
                phenotype="likely susceptible",
                comment=None,
            ),
            Namespace(
                mutation="V36A",
                section=expected_section,
                phenotype="likely susceptible",
                comment=None,
            ),
            Namespace(
                mutation="T40A",
                section=expected_section,
                phenotype="resistance possible",
                comment="unreliable",
            ),
        ]
        reader = WorksheetReader(worksheets)

        entries = list(reader)

        self.assertEqual(expected_entries, entries)

    def test_strike_through(self):
        row_data = [
            ["", "Example", "<"],
            ["", "Phenotype", "Comment"],
            ["WT", "likely susceptible"],
            ["V36A", "likely susceptible", "strike through to ignore"],
            ["T40A", "resistance possible"],
        ]
        worksheets = [create_worksheet("NS3_GT1b", row_data)]
        cell = worksheets[0].cell(4, 1)
        font = copy(cell.font)
        font.strike = True
        cell.font = font
        expected_section = Namespace(drug_name="Example", sheet_name="NS3_GT1b")
        expected_entries = [
            Namespace(
                mutation="WT",
                section=expected_section,
                phenotype="likely susceptible",
                comment=None,
            ),
            Namespace(
                mutation="T40A",
                section=expected_section,
                phenotype="resistance possible",
                comment=None,
            ),
        ]

        self.assertReads(expected_entries, worksheets)


class RulesWriterTest(TestCase):
    def setUp(self):
        self.entries = []
        self.phenotype_changes = []
        self.expected_rules = self.expected_errors = ""

    def assertWrites(self, expected_rules, entries):
        rules = StringIO()
        errors = StringIO()
        writer = RulesWriter(
            rules, errors, REFERENCES, phenotype_changes=self.phenotype_changes
        )

        writer.write(entries)
        writer.write_errors()

        self.assertEqual(self.expected_errors, errors.getvalue())
        self.assertEqual(expected_rules, rules.getvalue())

    def test_empty(self):
        """Empty spreadsheet generates an empty list of rules."""
        entries = []
        expected_rules = """\
[]
"""

        self.assertWrites(expected_rules, entries)

    def test_simple(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT1a")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="T40A", section=section, phenotype="resistance possible"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( T40A => 4 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_one_position_two_mutations(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT1a")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="T40A", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="T40R", section=section, phenotype="resistance possible"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( T40AR => 4 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_sofosbuvir_genotype1(self):
        self.maxDiff = None
        section1a = Namespace(drug_name="Sofosbuvir", sheet_name="NS5b_GT1a")
        section1b = Namespace(drug_name="Sofosbuvir", sheet_name="NS5b_GT1b")
        entries = [
            Namespace(mutation="WT", section=section1a, phenotype="likely susceptible"),
            Namespace(
                mutation="L159F", section=section1a, phenotype="resistance possible"
            ),
            Namespace(
                mutation="S282R", section=section1a, phenotype="resistance likely"
            ),
            Namespace(mutation="WT", section=section1b, phenotype="likely susceptible"),
            Namespace(
                mutation="S282R", section=section1b, phenotype="resistance likely"
            ),
        ]
        expected_rules = """\
- code: SOF-EPC
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS5b
    region: NS5b
    rules: SCORE FROM ( L159F => 4, S282R => 8 )
  - genotype: 1B
    reference: HCV1B-Con1-NS5b
    region: NS5b
    rules: SCORE FROM ( S282R => 8 )
  name: Sofosbuvir in Epclusa
- code: SOF-HAR
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS5b
    region: NS5b
    rules: SCORE FROM ( L159F => 4, S282R => 8 )
  - genotype: 1B
    reference: HCV1B-Con1-NS5b
    region: NS5b
    rules: SCORE FROM ( S282R => 8 )
  name: Sofosbuvir in Harvoni
"""

        self.assertWrites(expected_rules, entries)

    def test_sofosbuvir_genotype2(self):
        section = Namespace(drug_name="Sofosbuvir", sheet_name="NS5b_GT2")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(mutation="S282R", section=section, phenotype="resistance likely"),
        ]
        expected_rules = """\
- code: SOF-EPC
  genotypes:
  - genotype: '2'
    reference: HCV2-JFH-1-NS5b
    region: NS5b
    rules: SCORE FROM ( S282R => 8 )
  name: Sofosbuvir in Epclusa
- code: SOF-HAR
  genotypes:
  - genotype: '2'
    reference: HCV2-JFH-1-NS5b
    region: NS5b
    rules: SCORE FROM ( TRUE => "Not indicated" )
  name: Sofosbuvir in Harvoni
"""

        self.assertWrites(expected_rules, entries)

    def test_sofosbuvir_genotype6(self):
        self.maxDiff = None
        section = Namespace(drug_name="Sofosbuvir", sheet_name="NS5b_GT6")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(mutation="S282R", section=section, phenotype="resistance likely"),
        ]
        expected_rules = """\
- code: SOF-EPC
  genotypes:
  - genotype: '6'
    reference: HCV6-EUHK2-NS5b
    region: NS5b
    rules: SCORE FROM ( S282R => 8 )
  - genotype: 6E
    reference: HCV6-EUHK2-NS5b
    region: NS5b
    rules: SCORE FROM ( TRUE => "Not available" )
  name: Sofosbuvir in Epclusa
- code: SOF-HAR
  genotypes:
  - genotype: '6'
    reference: HCV6-EUHK2-NS5b
    region: NS5b
    rules: SCORE FROM ( TRUE => "Not indicated" )
  name: Sofosbuvir in Harvoni
"""

        self.assertWrites(expected_rules, entries)

    def test_sofosbuvir_282(self):
        self.maxDiff = None
        fold_range = Range(
            "<20 FS, likely susceptible",
            ">100 FS, resistance likely",
            changes=self.phenotype_changes,
        )
        section = Namespace(
            drug_name="Sofosbuvir",
            sheet_name="NS5b_GT1a",
            fold_range=fold_range,
            range_range=fold_range,
        )
        entries = [
            Namespace(
                mutation="WT",
                section=section,
                fold_shift="1x",
                clinical_ras=None,
                phenotype="likely susceptible",
            ),
            Namespace(
                mutation="S282R",
                section=section,
                fold_shift=None,
                clinical_ras="Yes",
                phenotype="resistance likely",
            ),
        ]
        expected_rules = """\
- code: SOF-EPC
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS5b
    region: NS5b
    rules: SCORE FROM ( S282R => 8 )
  name: Sofosbuvir in Epclusa
- code: SOF-HAR
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS5b
    region: NS5b
    rules: SCORE FROM ( S282R => 8 )
  name: Sofosbuvir in Harvoni
"""

        self.assertWrites(expected_rules, entries)

    def test_sofosbuvir_282_not_resistant(self):
        self.maxDiff = None
        fold_range = Range(
            "<20 FS, likely susceptible",
            ">100 FS, resistance likely",
            changes=self.phenotype_changes,
        )
        section = Namespace(
            drug_name="Sofosbuvir",
            sheet_name="NS5b_GT1a",
            fold_range=fold_range,
            range_range=fold_range,
        )
        entries = [
            Namespace(
                mutation="WT",
                section=section,
                fold_shift="1x",
                clinical_ras=None,
                phenotype="likely susceptible",
            ),
            Namespace(
                mutation="S282R",
                section=section,
                fold_shift=None,
                clinical_ras="Yes",
                phenotype="likely susceptible",
            ),
        ]
        expected_rules = """\
- code: SOF-EPC
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS5b
    region: NS5b
    rules: SCORE FROM ( TRUE => 0 )
  name: Sofosbuvir in Epclusa
- code: SOF-HAR
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS5b
    region: NS5b
    rules: SCORE FROM ( TRUE => 0 )
  name: Sofosbuvir in Harvoni
"""
        self.expected_errors = """\
Phenotype change: S282R (special case resistance likely) \
but is likely susceptible.
"""

        self.assertWrites(expected_rules, entries)

    def test_no_resistance(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT1a")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible")
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( TRUE => 0 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_not_indicated(self):
        section = Namespace(
            drug_name="Paritaprevir", sheet_name="NS3_GT1a", not_indicated=True
        )
        entries = [Namespace(section=section)]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( TRUE => "Not indicated" )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_missing_genotypes(self):
        section1 = Namespace(drug_name="Grazoprevir", sheet_name="NS3_GT1a")
        section2 = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT2")
        entries = [
            Namespace(mutation="WT", section=section1, phenotype="likely susceptible"),
            Namespace(mutation="WT", section=section2, phenotype="likely susceptible"),
        ]
        expected_rules = """\
- code: GZR
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( TRUE => 0 )
  - genotype: '2'
    reference: HCV2-JFH-1-NS3
    region: NS3
    rules: SCORE FROM ( TRUE => "Not indicated" )
  name: Grazoprevir
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( TRUE => "Not indicated" )
  - genotype: '2'
    reference: HCV2-JFH-1-NS3
    region: NS3
    rules: SCORE FROM ( TRUE => 0 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_missing_genotypes_with_subtypes(self):
        section1 = Namespace(drug_name="Grazoprevir", sheet_name="NS3_GT1a")
        section2 = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT2")
        entries = [
            Namespace(mutation="WT", section=section1, phenotype="likely susceptible"),
            Namespace(
                mutation="E30A", section=section1, phenotype="resistance possible"
            ),
            Namespace(mutation="WT", section=section2, phenotype="likely susceptible"),
            Namespace(
                mutation="S40W (GT2a)",
                section=section2,
                phenotype="resistance possible",
            ),
            Namespace(
                mutation="S40W (GT2b)", section=section2, phenotype="resistance likely"
            ),
        ]
        expected_rules = """\
- code: GZR
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( E30A => 4 )
  - genotype: '2'
    reference: HCV2-JFH-1-NS3
    region: NS3
    rules: SCORE FROM ( TRUE => "Not indicated" )
  name: Grazoprevir
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( TRUE => "Not indicated" )
  - genotype: '2'
    reference: HCV2-JFH-1-NS3
    region: NS3
    rules: SCORE FROM ( S40W => 4, S40W => 8 )
  name: Paritaprevir
"""
        self.expected_errors = """\
Conflicting phenotype: Paritaprevir in NS3_GT2:\
 S40W resistance possible => S40W resistance likely.
"""

        self.assertWrites(expected_rules, entries)

    def test_missing_genotypes_with_some_subtypes(self):
        section1 = Namespace(drug_name="Grazoprevir", sheet_name="NS3_GT2")
        section2 = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT2")
        entries = [
            Namespace(mutation="WT", section=section1, phenotype="likely susceptible"),
            Namespace(
                mutation="A30E", section=section1, phenotype="resistance possible"
            ),
            Namespace(mutation="WT", section=section2, phenotype="likely susceptible"),
            Namespace(
                mutation="S40W (GT2a)",
                section=section2,
                phenotype="resistance possible",
            ),
        ]
        expected_rules = """\
- code: GZR
  genotypes:
  - genotype: '2'
    reference: HCV2-JFH-1-NS3
    region: NS3
    rules: SCORE FROM ( A30E => 4 )
  name: Grazoprevir
- code: PTV
  genotypes:
  - genotype: '2'
    reference: HCV2-JFH-1-NS3
    region: NS3
    rules: SCORE FROM ( S40W => 4 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_exclude_zeroes(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT1a")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(mutation="S20A", section=section, phenotype="likely susceptible"),
            Namespace(mutation="T40A", section=section, phenotype="resistance likely"),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( T40A => 8 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_phenotype_case_insensitive(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT1a")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="T40A", section=section, phenotype="resiSTANce possible"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( T40A => 4 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_phenotype_invalid(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT1a")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(mutation="R20S", section=section, phenotype="bogus resistance"),
            Namespace(
                mutation="T40A", section=section, phenotype="resistance possible"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( T40A => 4 )
  name: Paritaprevir
"""
        self.expected_errors = """\
Invalid phenotype: NS3_GT1a, Paritaprevir, R20S: bogus resistance.
"""

        self.assertWrites(expected_rules, entries)

    def test_unknown_drug(self):
        section1 = Namespace(drug_name="Paulrevir", sheet_name="NS3_GT1a")
        section2 = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT1a")
        entries = [
            Namespace(mutation="WT", section=section1, phenotype="likely susceptible"),
            Namespace(mutation="T41R", section=section1, phenotype="resistance likely"),
            Namespace(mutation="WT", section=section2, phenotype="likely susceptible"),
            Namespace(
                mutation="T40A", section=section2, phenotype="resistance possible"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( T40A => 4 )
  name: Paritaprevir
"""
        self.expected_errors = """\
Unknown drug: NS3_GT1a: Paulrevir.
"""

        self.assertWrites(expected_rules, entries)

    def test_unknown_drugs(self):
        section1 = Namespace(drug_name="Paulrevir", sheet_name="NS3_GT1a")
        section2 = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT1a")
        section3 = Namespace(drug_name="Saynomorevir", sheet_name="NS3_GT1a")
        entries = [
            Namespace(mutation="WT", section=section1, phenotype="likely susceptible"),
            Namespace(mutation="T41R", section=section1, phenotype="resistance likely"),
            Namespace(mutation="WT", section=section2, phenotype="likely susceptible"),
            Namespace(
                mutation="T40A", section=section2, phenotype="resistance possible"
            ),
            Namespace(mutation="WT", section=section3, phenotype="likely susceptible"),
            Namespace(
                mutation="T20A", section=section3, phenotype="resistance possible"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( T40A => 4 )
  name: Paritaprevir
"""
        self.expected_errors = """\
Unknown drugs:
  NS3_GT1a: Paulrevir
  NS3_GT1a: Saynomorevir
"""

        self.assertWrites(expected_rules, entries)

    def test_wild_type_resistant(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT1a")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="resistance possible"),
            Namespace(
                mutation="T40A", section=section, phenotype="resistance possible"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( TRUE => 4, T40A => 4 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_wild_type_mismatch(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT3")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(mutation="A156G", section=section, phenotype="resistance likely"),
            Namespace(
                mutation="Q186L", section=section, phenotype="resistance possible"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: '3'
    reference: HCV3-S52-NS3
    region: NS3
    rules: SCORE FROM ( A156G => 8, D186L => 4 )
  name: Paritaprevir
"""
        self.expected_errors = """\
Mismatched wild type: NS3_GT3: Q186L in Paritaprevir expected D.
"""

        self.assertWrites(expected_rules, entries)

    def test_conflict_postponed(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT3")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="Q186L [Conflicting WT]",
                section=section,
                phenotype="resistance possible",
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: '3'
    reference: HCV3-S52-NS3
    region: NS3
    rules: SCORE FROM ( TRUE => 0 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_wildtype_conflict(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT3")
        entries = [
            Namespace(mutation="WT 1", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="WT 2 [Conflicting WT]",
                section=section,
                phenotype="resistance possible",
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: '3'
    reference: HCV3-S52-NS3
    region: NS3
    rules: SCORE FROM ( TRUE => 0 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_use_in_algorithm(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT3")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="D186L [Use in algorithm]",
                section=section,
                phenotype="resistance possible",
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: '3'
    reference: HCV3-S52-NS3
    region: NS3
    rules: SCORE FROM ( D186L => 4 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_wild_type_checked_with_subtype(self):
        section = Namespace(drug_name="Velpatasvir", sheet_name="NS5A_GT2")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="L28F (GT2b)", section=section, phenotype="resistance possible"
            ),
        ]
        expected_rules = """\
- code: VEL
  genotypes:
  - genotype: '2'
    reference: HCV2-JFH-1-NS5a
    region: NS5a
    rules: SCORE FROM ( F28F => 4 )
  name: Velpatasvir
"""
        self.expected_errors = """\
Mismatched wild type: NS5A_GT2: L28F in Velpatasvir expected F.
"""

        self.assertWrites(expected_rules, entries)

    def test_invalid_mutation(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT1a")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="T30A?", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="T40A", section=section, phenotype="resistance possible"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( T40A => 4 )
  name: Paritaprevir
"""
        self.expected_errors = """\
Invalid mutation: NS3_GT1a: T30A? (MutationSet text expects wild type \
(optional), position, and one or more variants.).
"""

        self.assertWrites(expected_rules, entries)

    def test_invalid_mutations(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT1a")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="T30A?", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="T30W?", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="T40A", section=section, phenotype="resistance possible"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( T40A => 4 )
  name: Paritaprevir
"""
        self.expected_errors = """\
Invalid mutations:
  NS3_GT1a: T30A? (MutationSet text expects wild type \
(optional), position, and one or more variants.)
  NS3_GT1a: T30W? (MutationSet text expects wild type \
(optional), position, and one or more variants.)
"""

        self.assertWrites(expected_rules, entries)

    def test_mutation_subtypes(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT2")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="A30E", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="S40W (GT2a)", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="S40W (GT2b)", section=section, phenotype="resistance likely"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: '2'
    reference: HCV2-JFH-1-NS3
    region: NS3
    rules: SCORE FROM ( A30E => 4, S40W => 4, S40W => 8 )
  name: Paritaprevir
"""
        self.expected_errors = """\
Conflicting phenotype: Paritaprevir in NS3_GT2:\
 S40W resistance possible => S40W resistance likely.
"""

        self.assertWrites(expected_rules, entries)

    def test_phenotype_conflict_susceptible(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT2")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="A30E", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="S40W (GT2a)", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="S40W (GT2b)", section=section, phenotype="likely susceptible"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: '2'
    reference: HCV2-JFH-1-NS3
    region: NS3
    rules: SCORE FROM ( A30E => 4, S40W => 4 )
  name: Paritaprevir
"""
        self.expected_errors = """\
Conflicting phenotype: Paritaprevir in NS3_GT2:\
 S40W likely susceptible => S40W resistance possible.
"""

        self.assertWrites(expected_rules, entries)

    def test_subtype_with_notes(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT2")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="A30E", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="S40W (GT2b_Added_Notes)",
                section=section,
                phenotype="resistance possible",
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: '2'
    reference: HCV2-JFH-1-NS3
    region: NS3
    rules: SCORE FROM ( A30E => 4, S40W => 4 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_subtypes_at_same_position(self):
        """Subtype mutations at the same position as shared mutations."""
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT2")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="S40E", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="S40W (GT2a)", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="S40W (GT2b)", section=section, phenotype="resistance likely"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: '2'
    reference: HCV2-JFH-1-NS3
    region: NS3
    rules: SCORE FROM ( S40W => 8, S40EW => 4 )
  name: Paritaprevir
"""
        self.expected_errors = """\
Conflicting phenotype: Paritaprevir in NS3_GT2:\
 S40EW resistance possible => S40W resistance likely.
"""

        self.assertWrites(expected_rules, entries)

    def test_redundant_subtype(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT1a")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="E30A", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="T40W (GT1a)", section=section, phenotype="resistance likely"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( E30A => 4, T40W => 8 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_identical_subtypes(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT2")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="S40W (GT2a)", section=section, phenotype="resistance likely"
            ),
            Namespace(
                mutation="S40W (GT2b)", section=section, phenotype="resistance likely"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: '2'
    reference: HCV2-JFH-1-NS3
    region: NS3
    rules: SCORE FROM ( S40W => 8 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_mismatched_subtype(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT3")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="S20A", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="T40W (GT2a)", section=section, phenotype="resistance likely"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: '3'
    reference: HCV3-S52-NS3
    region: NS3
    rules: SCORE FROM ( S20A => 4, T40W => 8 )
  name: Paritaprevir
"""
        self.expected_errors = """\
Invalid mutation: NS3_GT3: T40W (GT2a) (Mismatched subtype.).
"""

        self.assertWrites(expected_rules, entries)

    def test_combination_matches(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT1a")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="E30A", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="T40W", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="E30A+T40W", section=section, phenotype="resistance likely"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( E30A => 4, T40W => 4 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_combination_mixture(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT1a")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="E30A", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="E30W", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="E30A +E30W", section=section, phenotype="resistance possible"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( E30AW => 4 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_combination_change(self):
        """Combination score differs from parts, just report it for now."""
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT1a")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="E30A", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="E30A+R40W", section=section, phenotype="resistance likely"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( E30A => 4 )
  name: Paritaprevir
"""
        self.expected_errors = """\
Combination change: NS3_GT1a: E30A+R40W: 4 => 8.
"""

        self.assertWrites(expected_rules, entries)

    def test_combination_with_subtypes(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT2")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="S20L (GT2a)", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="S20L (GT2b)", section=section, phenotype="resistance likely"
            ),
            Namespace(
                mutation="S40W", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="E30A+S40W", section=section, phenotype="resistance possible"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: '2'
    reference: HCV2-JFH-1-NS3
    region: NS3
    rules: SCORE FROM ( S20L => 4, S20L => 8, S40W => 4 )
  name: Paritaprevir
"""
        self.expected_errors = """\
Conflicting phenotype: Paritaprevir in NS3_GT2:\
 S20L resistance possible => S20L resistance likely.
"""

        self.assertWrites(expected_rules, entries)

    def test_invalid_combination(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT1a")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="E30A", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="E30A+R40W@", section=section, phenotype="resistance likely"
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( E30A => 4 )
  name: Paritaprevir
"""
        self.expected_errors = """\
Invalid mutation: NS3_GT1a: E30A+R40W@ (MutationSet text expects wild type \
(optional), position, and one or more variants.).
"""

        self.assertWrites(expected_rules, entries)

    def test_combination_wildtype_override(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT1a")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="E30A", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="T40W", section=section, phenotype="resistance possible"
            ),
            Namespace(
                mutation="L30A+T40W [Conflicting WT]",
                section=section,
                phenotype="resistance likely",
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( E30A => 4, T40W => 4 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_monitored_positions(self):
        section = Namespace(
            drug_name="Paritaprevir", sheet_name="NS3_GT1a", monitored_positions=[80]
        )
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="Q80K", section=section, phenotype="resistance possible"
            ),
            Namespace(mutation="Q80L", section=section, phenotype="resistance likely"),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( Q80K => 4, Q80L => 8, Q80!KLQ => "Effect unknown" )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_monitored_positions_long(self):
        section = Namespace(
            drug_name="Paritaprevir",
            sheet_name="NS3_GT1a",
            monitored_positions=[80, 81],
        )
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="Q80K", section=section, phenotype="resistance possible"
            ),
            Namespace(mutation="Q80L", section=section, phenotype="resistance likely"),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( Q80K => 4, Q80L => 8, Q80!KLQ => "Effect unknown",
       D81!D => "Effect unknown" )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_monitored_position_invalid(self):
        section = Namespace(
            drug_name="Paritaprevir", sheet_name="NS3_GT1a", monitored_positions=[800]
        )
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="Q80K", section=section, phenotype="resistance possible"
            ),
            Namespace(mutation="Q80L", section=section, phenotype="resistance likely"),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( Q80K => 4, Q80L => 8 )
  name: Paritaprevir
"""
        self.expected_errors = """\
Invalid monitored position: NS3_GT1a: Paritaprevir 800 (max 631).
"""

        self.assertWrites(expected_rules, entries)

    def test_mutation_invalid_position(self):
        section = Namespace(drug_name="Paritaprevir", sheet_name="NS3_GT1a")
        entries = [
            Namespace(mutation="WT", section=section, phenotype="likely susceptible"),
            Namespace(
                mutation="Q80K", section=section, phenotype="resistance possible"
            ),
            Namespace(mutation="Q800L", section=section, phenotype="resistance likely"),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( Q80K => 4 )
  name: Paritaprevir
"""
        self.expected_errors = """\
Invalid mutation position: NS3_GT1a: Paritaprevir Q800L (max 631).
"""

        self.assertWrites(expected_rules, entries)

    def test_fold_shifts_match(self):
        fold_range = Range(
            "<20 FS, likely susceptible",
            ">100 FS, resistance likely",
            changes=self.phenotype_changes,
        )
        section = Namespace(
            drug_name="Paritaprevir",
            sheet_name="NS3_GT1a",
            fold_range=fold_range,
            range_range=fold_range,
        )
        entries = [
            Namespace(
                mutation="WT",
                section=section,
                phenotype="likely susceptible",
                fold_shift="1x",
                clinical_ras=None,
            ),
            Namespace(
                mutation="Q80K",
                section=section,
                phenotype="resistance possible",
                fold_shift="50x",
                clinical_ras=None,
            ),
            Namespace(
                mutation="Q80L",
                section=section,
                phenotype="resistance likely",
                fold_shift="200x",
                clinical_ras=None,
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( Q80K => 4, Q80L => 8 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_fold_shifts_without_clinical(self):
        fold_range = Range(
            "<20 FS, likely susceptible",
            ">100 FS, resistance likely",
            changes=self.phenotype_changes,
        )
        section = Namespace(
            drug_name="Paritaprevir",
            sheet_name="NS3_GT1a",
            fold_range=fold_range,
            range_range=fold_range,
        )
        entries = [
            Namespace(
                mutation="WT",
                section=section,
                phenotype="likely susceptible",
                fold_shift="1x",
            ),
            Namespace(
                mutation="Q80K",
                section=section,
                phenotype="resistance possible",
                fold_shift="50x",
            ),
            Namespace(
                mutation="Q80L",
                section=section,
                phenotype="resistance likely",
                fold_shift="200x",
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( Q80K => 4, Q80L => 8 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_fold_shifts_differ(self):
        fold_range = Range(
            "<20 FS, likely susceptible",
            ">100 FS, resistance likely",
            "NS3_GT1a: Paritaprevir",
            self.phenotype_changes,
        )
        section = Namespace(
            drug_name="Paritaprevir",
            sheet_name="NS3_GT1a",
            fold_range=fold_range,
            range_range=fold_range,
        )
        entries = [
            Namespace(
                mutation="WT",
                section=section,
                phenotype="likely susceptible",
                fold_shift="1x",
                clinical_ras=None,
            ),
            Namespace(
                mutation="Q80K",
                section=section,
                phenotype="resistance possible",
                fold_shift="2x",
                clinical_ras=None,
            ),
            Namespace(
                mutation="Q80L",
                section=section,
                phenotype="resistance likely",
                fold_shift="100x",  # Should be > 100.0
                clinical_ras=None,
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( Q80K => 4, Q80L => 8 )
  name: Paritaprevir
"""
        self.expected_errors = """\
Phenotype changes:
  NS3_GT1a: Paritaprevir Q80K (2x/no in <20, >100 likely susceptible) \
but is resistance possible
  NS3_GT1a: Paritaprevir Q80L (100x/no in <20, >100 resistance possible) \
but is resistance likely
"""

        self.assertWrites(expected_rules, entries)

    def test_fold_shift_no_levels(self):
        section = Namespace(
            drug_name="Paritaprevir",
            sheet_name="NS3_GT1a",
            fold_range=None,
            range_range=None,
        )
        entries = [
            Namespace(
                mutation="WT",
                section=section,
                phenotype="likely susceptible",
                fold_shift="1x",
                clinical_ras=None,
            ),
            Namespace(
                mutation="Q80K",
                section=section,
                phenotype="resistance possible",
                fold_shift="2x",
                clinical_ras=None,
            ),
            Namespace(
                mutation="Q80L",
                section=section,
                phenotype="resistance likely",
                fold_shift="100x",
                clinical_ras=None,
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( Q80K => 4, Q80L => 8 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_fold_shift_ranges(self):
        fold_range = Range(
            "<10 FS, likely susceptible",
            ">100 FS, resistance likely",
            "NS3_GT1a: Paritaprevir",
            self.phenotype_changes,
        )
        range_range = Range(
            "<=50 FS, likely susceptible",
            ">=100 FS, resistance likely",
            "NS3_GT1a: Paritaprevir",
            self.phenotype_changes,
        )
        section = Namespace(
            drug_name="Paritaprevir",
            sheet_name="NS3_GT1a",
            fold_range=fold_range,
            range_range=range_range,
        )
        entries = [
            Namespace(
                mutation="WT",
                section=section,
                phenotype="likely susceptible",
                fold_shift="1x",
                clinical_ras=None,
            ),
            Namespace(
                mutation="Q80K",
                section=section,
                phenotype="likely susceptible",
                fold_shift="2-50x",
                clinical_ras=None,
            ),
            Namespace(
                mutation="Q80L",
                section=section,
                phenotype="resistance likely",
                fold_shift="100-1000x",
                clinical_ras=None,
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( Q80L => 8 )
  name: Paritaprevir
"""

        self.assertWrites(expected_rules, entries)

    def test_fold_shift_int(self):
        fold_range = Range(
            "<20 FS, likely susceptible",
            ">100 FS, resistance likely",
            "NS3_GT1a: Paritaprevir",
            self.phenotype_changes,
        )
        section = Namespace(
            drug_name="Paritaprevir",
            sheet_name="NS3_GT1a",
            fold_range=fold_range,
            range_range=fold_range,
        )
        entries = [
            Namespace(
                mutation="WT",
                section=section,
                phenotype="likely susceptible",
                fold_shift=1,
                clinical_ras=None,
            ),
            Namespace(
                mutation="Q80K",
                section=section,
                phenotype="resistance possible",
                fold_shift=2,
                clinical_ras=None,
            ),
            Namespace(
                mutation="Q80L",
                section=section,
                phenotype="resistance likely",
                fold_shift=100,  # Should be > 100.0
                clinical_ras=None,
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( Q80K => 4, Q80L => 8 )
  name: Paritaprevir
"""
        self.expected_errors = """\
Phenotype changes:
  NS3_GT1a: Paritaprevir Q80K (2/no in <20, >100 likely susceptible) \
but is resistance possible
  NS3_GT1a: Paritaprevir Q80L (100/no in <20, >100 resistance possible) \
but is resistance likely
"""

        self.assertWrites(expected_rules, entries)

    def test_fold_shift_missing(self):
        fold_range = Range(
            "<20 FS, likely susceptible",
            ">100 FS, resistance likely",
            "NS3_GT1a: Paritaprevir",
            self.phenotype_changes,
        )
        section = Namespace(
            drug_name="Paritaprevir",
            sheet_name="NS3_GT1a",
            fold_range=fold_range,
            range_range=fold_range,
        )
        entries = [
            Namespace(
                mutation="WT",
                section=section,
                phenotype="likely susceptible",
                fold_shift="1x",
                clinical_ras=None,
            ),
            Namespace(
                mutation="Q80K",
                section=section,
                phenotype="resistance possible",
                fold_shift=None,
                clinical_ras=None,
            ),
            Namespace(
                mutation="G90L",
                section=section,
                phenotype="resistance likely",
                fold_shift=None,
                clinical_ras=None,
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( Q80K => 4, G90L => 8 )
  name: Paritaprevir
"""
        self.expected_errors = """\
Missing fold shift: NS3_GT1a: Paritaprevir Q80K, G90L.
"""

        self.assertWrites(expected_rules, entries)

    def test_clinical_without_fold_shift(self):
        fold_range = Range(
            "<20 FS, likely susceptible",
            ">100 FS, resistance likely",
            "NS3_GT1a: Paritaprevir",
            self.phenotype_changes,
        )
        section = Namespace(
            drug_name="Paritaprevir",
            sheet_name="NS3_GT1a",
            fold_range=fold_range,
            range_range=fold_range,
        )
        entries = [
            Namespace(
                mutation="WT",
                section=section,
                phenotype="likely susceptible",
                fold_shift="1x",
                clinical_ras=None,
            ),
            Namespace(
                mutation="Q80K",
                section=section,
                phenotype="resistance possible",
                fold_shift=None,
                clinical_ras="yes",
            ),
            Namespace(
                mutation="G90L",
                section=section,
                phenotype="resistance likely",
                fold_shift=None,
                clinical_ras=None,
            ),
        ]
        expected_rules = """\
- code: PTV
  genotypes:
  - genotype: 1A
    reference: HCV1A-H77-NS3
    region: NS3
    rules: SCORE FROM ( Q80K => 4, G90L => 8 )
  name: Paritaprevir
"""
        self.expected_errors = """\
Missing fold shift: NS3_GT1a: Paritaprevir G90L.
"""

        self.assertWrites(expected_rules, entries)
