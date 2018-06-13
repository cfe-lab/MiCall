import operator
from argparse import ArgumentParser, ArgumentDefaultsHelpFormatter, FileType, Namespace
from collections import namedtuple, defaultdict
from copy import deepcopy
from functools import partial
import itertools
from itertools import product, groupby
from operator import itemgetter, attrgetter
import os
import re
import sys

import yaml
from pyvdrm.hcvr import HCVR
from pyvdrm.vcf import MutationSet, VariantCalls
from openpyxl import load_workbook

from micall.core.project_config import ProjectConfig

READY_TABS = ('NS3_GT1a',
              'NS3_GT1b',
              'NS3_GT2',
              'NS3_GT3',
              'NS3_GT4',
              'NS3_GT5',
              'NS3_GT6',
              'NS5A_GT1a',
              'NS5A_GT1b',
              'NS5A_GT2',
              'NS5A_GT3',
              'NS5A_GT4',
              'NS5A_GT5',
              'NS5A_GT6',
              'NS5B_GT1a',
              'NS5B_GT1b',
              'NS5B_GT2',
              'NS5B_GT3',
              'NS5B_GT4',
              'NS5B_GT5',
              'NS5B_GT6')
PHENOTYPE_SCORES = {'likely susceptible': 0,
                    'resistance possible': 4,
                    'resistance likely': 8,
                    'effect unknown': 'effect unknown'}
HCV_REGIONS = ('NS3', 'NS5a', 'NS5b')
SPLIT_GENOTYPES = {'2': ('2a', '2b'), '4': ('4a', '4d')}
OBSOLETE_DRUGS = ('Boceprevir', 'Telaprevir')
RuleSet = namedtuple(
    'RuleSet',
    ['region',
     'genotype',
     'drug_name',
     'first_column',
     'phenotype_column',
     'last_column',
     'mutations',  # {condition: score}
     'fold_shifts'])  # {condition: shift_text}
RuleSet.__new__.__defaults__ = (None, )
Reference = namedtuple('Reference', 'name sequence')


def parse_args():
    parser = ArgumentParser(description='Read HCV rules from a spreadsheet.',
                            formatter_class=ArgumentDefaultsHelpFormatter)
    parser.add_argument('spreadsheet',
                        type=partial(load_workbook, read_only=False))
    parser.add_argument('rules_yaml',
                        type=FileType('w'),
                        help='YAML file to write new rules to')
    return parser.parse_args()


class Range:
    lower_pattern = \
        r'(([</=]*) *([\d.]+)(-([\d.]+))?)[x ]*FS( range)?, likely susceptibi?le$'
    upper_pattern = \
        r'(([>/=]*) *([\d.]+)(-([\d.]+))?)[x ]*FS( range)?, resistance likely$'
    score_phenotypes = {
        score: phenotype
        for phenotype, score in PHENOTYPE_SCORES.items()}

    def __init__(self,
                 lower_text,
                 upper_text,
                 drug=None,
                 changes=None,
                 invalid_fold_shifts=None):
        self.lower_original = lower_text
        self.upper_original = upper_text
        self.drug = drug
        self.changes = changes if changes is not None else []
        self.invalid_fold_shifts = (invalid_fold_shifts
                                    if invalid_fold_shifts is not None
                                    else [])
        lower_match = re.match(self.lower_pattern,
                               lower_text,
                               flags=re.IGNORECASE)
        if not lower_match:
            self.lower = self.lower_operator = None
        else:
            self.lower_text = lower_match.group(1)
            self.lower = float(lower_match.group(5)
                               or lower_match.group(3))
            lower_operator_text = lower_match.group(2)
            if lower_operator_text == '<':
                self.lower_operator = operator.lt
            elif lower_operator_text in ('<=', '</='):
                self.lower_operator = operator.le
            elif lower_match.group(5) and lower_operator_text == '':
                self.lower_operator = operator.le
            else:
                self.lower = None
        upper_match = re.match(self.upper_pattern,
                               upper_text,
                               flags=re.IGNORECASE)
        if not upper_match:
            self.upper = self.upper_operator = None
        else:
            self.upper_text = upper_match.group(1)
            self.upper = float(upper_match.group(3))
            upper_operator_text = upper_match.group(2)
            if upper_operator_text == '>':
                self.upper_operator = operator.gt
            elif upper_operator_text in ('>=', '>/='):
                self.upper_operator = operator.ge
            elif upper_match.group(5) and upper_operator_text == '':
                self.upper_operator = operator.ge
            else:
                self.upper = None
        self.is_valid = self.lower is not None and self.upper is not None
        if self.lower is None:
            self.invalid_fold_shifts.append(
                f'{self.drug} lower fold shift: {lower_text}')
        if self.upper is None:
            self.invalid_fold_shifts.append(
                f'{self.drug} upper fold shift: {upper_text}')

    def __repr__(self):
        return f'Range({self.lower_original!r}, {self.upper_original!r})'

    def __str__(self):
        return f'{self.lower_text}, {self.upper_text}'

    def __eq__(self, other: 'Range'):
        return (self.lower == other.lower and
                self.upper == other.upper and
                self.lower_operator == other.lower_operator and
                self.upper_operator == other.upper_operator)

    def validate_phenotype(self,
                           mutation,
                           phenotype,
                           fold_shift_text,
                           clinical_ras='no',
                           expected_phenotype=None):
        if expected_phenotype is not None:
            expectation_display = 'special case ' + expected_phenotype
        else:
            match = re.match(
                r'([<>/=]*) *(([\d.,]+)x?|([\d.,]+)x? *- *>? *([\d.,]+)x)$',
                fold_shift_text,
                flags=re.IGNORECASE)
            if match is None:
                self.invalid_fold_shifts.append(
                    f'{self.drug} {mutation} {fold_shift_text!r}')
                return
            if not self.is_valid:
                return
            if match.group(3):
                fold_shift = float(match.group(3).replace(',', ''))
                comparison = match.group(1)
                if comparison == '>':
                    fold_shift += 0.1
                elif comparison == '<':
                    fold_shift -= 0.1
                if self.lower_operator(fold_shift, self.lower):
                    expected_score = 0
                elif self.upper_operator(fold_shift, self.upper):
                    expected_score = 8
                else:
                    expected_score = 4
            else:
                # A range of fold shifts always allows overlap at boundaries.
                lower_fold_shift = float(match.group(4))
                upper_fold_shift = float(match.group(5))
                if upper_fold_shift <= self.lower:
                    expected_score = 0
                elif lower_fold_shift >= self.upper:
                    expected_score = 8
                elif not (lower_fold_shift < self.lower or
                          upper_fold_shift > self.upper):
                    expected_score = 4
                else:
                    expected_score = None
            if expected_score is None:
                expected_phenotype = 'ambiguous'
            else:
                if clinical_ras == 'yes':
                    expected_score = min(8, expected_score + 4)
                expected_phenotype = self.score_phenotypes[expected_score]
            expectation_display = (f'{fold_shift_text}/{clinical_ras} '
                                   f'in {self} {expected_phenotype}')
        if phenotype != expected_phenotype:
            prefix = self.drug and (self.drug + ' ') or ''
            self.changes.append(
                f'{prefix}{mutation} ({expectation_display}) but is {phenotype}')


class SplitDumper(yaml.SafeDumper):
    def write_plain(self, text, split=True):
        delimiter = ','
        if split:
            pieces = text.split(delimiter)
        else:
            pieces = [text]
        buffer = ''
        for i, piece in enumerate(pieces):
            if i > 0:
                buffer += delimiter
            if self.column-1 + len(buffer) + len(piece) <= self.best_width:
                buffer += piece
            else:
                super(SplitDumper, self).write_plain(buffer, split)
                self.write_indent()
                buffer = piece
        super(SplitDumper, self).write_plain(buffer)


class WorksheetReader:
    def __init__(self, worksheets, *footer_readers):
        self.worksheets = worksheets
        self.missing_wild_types = []
        self.missing_monitored_positions = []
        self.indicated_obsolete_drugs = []
        self.footer_readers = footer_readers

    def __iter__(self):
        for ws in self.worksheets:
            yield from self._iter_worksheet(ws)

    def _iter_worksheet(self, ws):
        wild_type_row, footer_row = self._find_row_limits(ws)
        if wild_type_row is None:
            self.missing_wild_types.append(ws.title)
            return
        header_row = wild_type_row - 1
        drug_row = wild_type_row - 2
        for cell_range in ws.merged_cells:
            if cell_range.min_row != drug_row:
                continue
            cell = ws.cell(cell_range.min_row, cell_range.min_col)
            if cell.value is None:
                continue
            drug_name, *_ = cell.value.split()
            column_headings = {
                col: self._format_heading(ws.cell(header_row, col).value)
                for col in range(cell_range.min_col, cell_range.max_col + 1)
                if ws.cell(header_row, col).value is not None}
            column_headings[1] = 'mutation'
            section = Namespace(drug_name=drug_name, sheet_name=ws.title)
            if not self._read_footer(ws, cell_range, footer_row, section):
                section.not_indicated = True
                if section.drug_name not in OBSOLETE_DRUGS:
                    yield Namespace(section=section)
                continue
            if section.drug_name in OBSOLETE_DRUGS:
                self.indicated_obsolete_drugs.append(
                    '{} in {}'.format(section.drug_name, section.sheet_name))
            for row_num in range(wild_type_row, footer_row):
                if ws.cell(row_num, 1).font.strike:
                    continue
                fields = {heading: ws.cell(row_num, col).value
                          for col, heading in column_headings.items()}
                for field in fields:
                    value = fields[field]
                    if value is not None:
                        fields[field] = str(value).strip()
                fields['mutation'] = re.sub(r'del$', 'd', fields['mutation'])
                fields['mutation'] = re.sub(r'ins$', 'i', fields['mutation'])
                entry = Namespace(section=section, **fields)
                if entry.phenotype is not None:
                    yield entry

    def _read_footer(self, ws, cell_range, footer_row, section):
        footer_coordinates = product(
            range(footer_row, ws.max_row + 1),
            range(cell_range.min_col, cell_range.max_col + 1))
        for row, col in footer_coordinates:
            value = str(ws.cell(row, col).value).lower().strip()
            if value == 'not indicated in canada':
                return False
        for footer_reader in self.footer_readers:
            footer_coordinates = product(
                range(footer_row, ws.max_row + 1),
                range(cell_range.min_col, cell_range.max_col + 1))
            footer_reader.read(ws, footer_coordinates, section)
        return True

    @staticmethod
    def _find_row_limits(ws):
        wild_type_row = None
        footer_row = ws.max_row + 1
        for row_num in range(1, ws.max_row + 1):
            mutation = ws.cell(row_num, 1).value
            if wild_type_row is None:
                if mutation and mutation.startswith('WT'):
                    wild_type_row = row_num
            else:
                if mutation is None:
                    footer_row = row_num
                    break
        return wild_type_row, footer_row

    @staticmethod
    def _format_heading(text):
        heading = text.lower()
        heading = re.sub(r'[^a-z0-9]', '_', heading)
        return heading.strip('_')

    def write_errors(self, report):
        if self.missing_wild_types:
            report.write('No wild type found in {}.\n'.format(
                ', '.join(self.missing_wild_types)))
        if self.indicated_obsolete_drugs:
            report.write('Obsolete drugs still indicated: {}.\n'.format(
                ', '.join(self.indicated_obsolete_drugs)))
        for footer_reader in self.footer_readers:
            footer_reader.write_errors(report)


class MonitoredPositionsReader:
    def __init__(self):
        self.missing = []

    def read(self, ws, footer_coordinates, section):
        monitored_positions = None
        for row_num, col_num in footer_coordinates:
            cell = ws.cell(row_num, col_num)
            label = cell.value and str(cell.value).lower()
            if label and label.startswith('positions monitored'):
                monitored_positions = ws.cell(row_num + 1, col_num).value
                break
        if monitored_positions is None:
            monitored_positions = []
            self.missing.append(
                '{} in {}'.format(section.drug_name, section.sheet_name))
        else:
            monitored_positions = str(monitored_positions)
            if monitored_positions.lower() == 'none':
                monitored_positions = []
            else:
                monitored_positions = list(map(
                    int,
                    str(monitored_positions).split(',')))
        section.monitored_positions = monitored_positions

    def write_errors(self, report):
        if self.missing:
            report.write('No monitored positions for {}.\n'.format(
                ', '.join(self.missing)))


class FoldRangesReader:
    def __init__(self, print_errors=True):
        self.print_errors = print_errors
        self.invalid = []
        self.changes = []
        self.lower_pattern = \
            r'([</=]*) *([\d.]+)(-([\d.]+))?[x ]*FS( range)?, likely susceptibi?le$'
        self.upper_pattern = \
            r'([>/=]*) *([\d.]+)(-([\d.]+))?[x ]*FS( range)?, resistance likely$'

    def read(self, ws, footer_coordinates, section):
        label_positions = {}  # {label: (row, col)}
        for row_num, col_num in footer_coordinates:
            label = ws.cell(row_num, col_num).value
            if not label:
                continue
            label = str(label).lower()
            label = label.replace('virto', 'vitro')
            label = label.replace('susecptibility', 'susceptibility')
            label = label.strip(' -:*')
            label_positions[label] = (row_num, col_num)
        section.fold_range = self.find_fold_range(
            'in vitro drug susceptibility',
            label_positions,
            section,
            ws)
        section.range_range = self.find_fold_range(
                'range fs values',
                label_positions,
                section,
                ws)
        if section.range_range is None:
            section.range_range = section.fold_range

    def find_fold_range(self, range_label, label_positions, section, ws):
        row_num, col_num = label_positions.get(range_label,
                                               (None, None))
        if row_num is None:
            return None
        lower_text = ws.cell(row_num + 1, col_num).value
        if str(lower_text).lower().strip(' -*') == 'absolute fs values':
            row_num += 1
            lower_text = ws.cell(row_num + 1, col_num).value
        upper_text = ws.cell(row_num + 3, col_num).value
        fold_range = Range(lower_text,
                           upper_text,
                           f'{section.drug_name} in {section.sheet_name}',
                           self.changes,
                           self.invalid)
        return fold_range

    def write_errors(self, report):
        if self.print_errors:
            for message in self.invalid:
                print(message, file=report)


class RulesWriter:
    def __init__(self,
                 rules_file,
                 errors_file,
                 references,
                 invalid_fold_shifts=None,
                 phenotype_changes=None):
        self.rules_file = rules_file
        self.errors_file = errors_file
        self.references = references
        self.unknown_drugs = []
        self.invalid_mutations = []
        self.invalid_mutation_positions = []
        self.combination_changes = []
        self.phenotype_conflicts = []
        self.invalid_fold_shifts = (invalid_fold_shifts
                                    if invalid_fold_shifts is not None
                                    else [])
        self.phenotype_changes = (phenotype_changes
                                  if phenotype_changes is not None
                                  else [])
        self.invalid_phenotypes = []
        self.invalid_positions = []
        self.bad_wild_types = []

        # {(sheet, drug_name): [mutation]}
        self.missing_fold_shifts = defaultdict(list)

        self.score_phenotypes = {
            score: phenotype
            for phenotype, score in PHENOTYPE_SCORES.items()}

    def write(self, entries):
        drug_codes = load_drug_codes()
        drug_summaries = {}
        for section, section_entries in groupby(entries, attrgetter('section')):
            try:
                drug_summary = self._find_drug_summary(drug_summaries,
                                                       drug_codes,
                                                       section)
            except KeyError:
                continue
            sheet_region, genotype = section.sheet_name.split('_GT')
            lower_region = sheet_region.lower()
            for known_region in HCV_REGIONS:
                if known_region.lower() == lower_region:
                    region = known_region
                    break
            else:
                region = sheet_region
            genotype = genotype.upper()
            reference = self.references[(genotype, region)]
            reference_name = reference.name
            # {(pos, genotype_override): {score: MutationSet}}
            score_map = defaultdict(dict)
            combinations = {}  # {mutation: score}
            if getattr(section, 'not_indicated', False):
                # noinspection PyTypeChecker
                score_map[None]['Not indicated'] = 'TRUE'
            else:
                for entry in section_entries:
                    self._score_mutation(entry,
                                         section,
                                         genotype,
                                         score_map,
                                         combinations,
                                         reference.sequence)

            for positions, subtype in self._expand_score_map(score_map,
                                                             genotype):
                self._check_combinations(combinations, positions, section)
                self._check_conflicts(positions, section)
                self._monitor_positions(section, positions, reference)
                score_formula = build_score_formula(positions)
                drug_summary['genotypes'].append(dict(genotype=subtype,
                                                      region=region,
                                                      reference=reference_name,
                                                      rules=score_formula))
        self._check_not_indicated(drug_summaries)
        self._override_sofosbuvir(drug_summaries)
        drugs = sorted(drug_summaries.values(), key=itemgetter('code'))
        yaml.dump(drugs, self.rules_file, default_flow_style=False, Dumper=SplitDumper)

    @staticmethod
    def _override_sofosbuvir(drug_summaries):
        drug_summary_harvoni = drug_summaries.pop('Sofosbuvir', None)
        if drug_summary_harvoni is None:
            return
        drug_summary_epclusa = deepcopy(drug_summary_harvoni)
        drug_summary_epclusa['code'] = 'SOF-EPC'
        drug_summary_epclusa['name'] = 'Sofosbuvir in Epclusa'
        for drug_genotype in list(drug_summary_epclusa['genotypes']):
            if drug_genotype['genotype'] == '6':
                extra_genotype = deepcopy(drug_genotype)
                extra_genotype['genotype'] = '6E'
                extra_genotype['rules'] = 'SCORE FROM ( TRUE => "Not available" )'
                drug_summary_epclusa['genotypes'].append(extra_genotype)
        drug_summaries[drug_summary_epclusa['name']] = drug_summary_epclusa
        drug_summary_harvoni['name'] = 'Sofosbuvir in Harvoni'
        for drug_genotype in list(drug_summary_harvoni['genotypes']):
            if get_main_genotype(drug_genotype['genotype']) == '1':
                pass
            else:
                drug_genotype['rules'] = 'SCORE FROM ( TRUE => "Not indicated" )'
        drug_summaries[drug_summary_harvoni['name']] = drug_summary_harvoni

    @staticmethod
    def _expand_score_map(score_map, main_genotype):
        # {genotype_override: {pos: {score: MutationSet}}}
        override_map = defaultdict(lambda: defaultdict(dict))
        base_positions = defaultdict(dict)  # {pos: {score: MutationSet}}
        for key, scores in score_map.items():
            if key is None:
                pos = genotype_override = None
            else:
                pos, genotype_override = key
            if genotype_override is None:
                base_positions[pos] = scores
            else:
                override_map[genotype_override][pos] = scores
        if not override_map:
            return [(base_positions, main_genotype)]
        subtypes = SPLIT_GENOTYPES[main_genotype]
        for genotype_override in subtypes:
            positions = override_map[genotype_override]
            for pos, base_scores in base_positions.items():
                for score, base_mutation_set in base_scores.items():
                    subtype_mutation_set = positions[pos].get(score)
                    if subtype_mutation_set is None:
                        combined_mutation_set = base_mutation_set
                    else:
                        combined_mutation_set = MutationSet(
                            mutations=base_mutation_set.mutations |
                            subtype_mutation_set.mutations)
                    positions[pos][score] = combined_mutation_set
        first_positions = override_map[subtypes[0]]
        if all(override_map[other_subtype] == first_positions
               for other_subtype in subtypes[1:]):
            return [(first_positions, main_genotype)]
        return [(positions, subtype.upper())
                for subtype, positions in override_map.items()]

    def _check_not_indicated(self, drug_summaries):
        all_genotypes = {get_unsplit_genotype(genotype['genotype'])
                         for drug_summary in drug_summaries.values()
                         for genotype in drug_summary['genotypes']}
        positions = {None: {'Not indicated': 'TRUE'}}
        score_formula = build_score_formula(positions)
        for drug_summary in drug_summaries.values():
            drug_genotypes = {get_unsplit_genotype(genotype['genotype'])
                              for genotype in drug_summary['genotypes']}
            missing_genotypes = all_genotypes - drug_genotypes
            main_genotypes = defaultdict(set)
            for genotype in missing_genotypes:
                genotype_group = main_genotypes[get_main_genotype(genotype)]
                genotype_group.add(genotype)

            drug_region, = {genotype['region']
                            for genotype in drug_summary['genotypes']}
            for main_genotype, subtypes in main_genotypes.items():
                if len(subtypes) > 1:
                    genotype_name = main_genotype
                else:
                    genotype_name, = subtypes
                try:
                    reference = self.references[(genotype_name, drug_region)]
                except KeyError:
                    if main_genotype != genotype_name:
                        reference = self.references[(genotype_name[:-1], drug_region)]
                    else:
                        raise
                drug_summary['genotypes'].append(dict(genotype=genotype_name,
                                                      region=drug_region,
                                                      reference=reference.name,
                                                      rules=score_formula))
            drug_summary['genotypes'].sort(key=itemgetter('genotype'))

    def _score_mutation(self,
                        entry,
                        section,
                        genotype,
                        score_map,
                        combinations,
                        reference):
        mutation = entry.mutation
        phenotype = entry.phenotype.lower()
        try:
            score = PHENOTYPE_SCORES[phenotype]
        except KeyError:
            self.invalid_phenotypes.append(
                '{}, {}, {}: {}'.format(section.sheet_name,
                                        section.drug_name,
                                        entry.mutation,
                                        entry.phenotype))
            return

        self._check_fold_shift(entry, section, phenotype)
        if mutation.startswith('WT'):
            # noinspection PyTypeChecker
            score_map[None][score] = 'TRUE'
            return

        match = re.match(r'([^(]*?) *(\(GT([^_]+).*\))? *(\[Conflicting WT\])?$',
                         mutation,
                         flags=re.IGNORECASE)
        if match is None:
            core_mutation = mutation
            genotype_override = None
            is_wild_type_checked = True
        else:
            core_mutation = match.group(1)
            genotype_override = match.group(3)
            is_wild_type_checked = match.group(4) is None
            if genotype_override is None:
                pass
            elif genotype_override.upper() == genotype:
                genotype_override = None
            elif get_main_genotype(genotype_override) != genotype:
                self.invalid_mutations.append(
                    '{}: {} (Mismatched subtype.)'.format(section.sheet_name,
                                                          mutation))
                genotype_override = None
            if genotype_override:
                genotype_override = None  # Ignoring overrides until issue #443.
        if '+' in core_mutation or ' ' in core_mutation:
            if not is_wild_type_checked:
                core_mutation = replace_wild_types(core_mutation, reference)
            combinations[core_mutation] = score
            return
        try:
            new_mutation_set = MutationSet(core_mutation)
        except ValueError as ex:
            self.invalid_mutations.append('{}: {} ({})'.format(
                section.sheet_name,
                core_mutation,
                ex))
            return
        max_pos = len(reference)
        if new_mutation_set.pos > max_pos:
            self.invalid_mutation_positions.append(
                f'{section.sheet_name}: {section.drug_name} '
                f'{new_mutation_set} (max {max_pos})')
            return
        expected_wild_type = reference[new_mutation_set.pos-1]
        if expected_wild_type != new_mutation_set.wildtype:
            if is_wild_type_checked:
                self.bad_wild_types.append(
                    f'{section.sheet_name}: {new_mutation_set} in '
                    f'{section.drug_name} expected {expected_wild_type}')
            new_mutation_set = MutationSet(
                expected_wild_type + str(new_mutation_set)[1:])
        pos_scores = score_map[(new_mutation_set.pos, genotype_override)]
        old_mutation_set = pos_scores.get(score)
        if old_mutation_set is not None:
            new_mutation_set = MutationSet(
                wildtype=old_mutation_set.wildtype,
                pos=old_mutation_set.pos,
                mutations=(old_mutation_set.mutations |
                           new_mutation_set.mutations))
        pos_scores[score] = new_mutation_set

    def _check_fold_shift(self, entry, section, phenotype):
        fold_range = getattr(section, 'fold_range', None)
        if fold_range is None:
            # Range wasn't found, so nothing to check.
            return
        clinical_ras = getattr(entry, 'clinical_ras', None)
        expected_phenotype = None
        if entry.fold_shift is not None:
            fold_shift_text = str(entry.fold_shift)
        else:
            if (section.drug_name == 'Sofosbuvir' and
                    entry.mutation.startswith('S282')):
                expected_phenotype = 'resistance likely'
                fold_shift_text = ''
            elif clinical_ras is not None:
                fold_shift_text = '1x'
            else:
                mutations = self.missing_fold_shifts[
                    (section.sheet_name, section.drug_name)]
                mutations.append(entry.mutation)
                return
        clinical_ras = clinical_ras and clinical_ras.lower() or 'no'
        if '-' in fold_shift_text:
            fold_range = section.range_range
        fold_range.validate_phenotype(entry.mutation,
                                      phenotype,
                                      fold_shift_text,
                                      clinical_ras,
                                      expected_phenotype)

    def _check_combinations(self, combinations, positions, section):
        for combination, combination_score in combinations.items():
            try:
                component_score = calculate_component_score(combination, positions)
            except ValueError as ex:
                self.invalid_mutations.append('{}: {} ({})'.format(
                    section.sheet_name,
                    combination,
                    ex))
                continue

            component_score = min(component_score, 8)
            if component_score != combination_score:
                self.combination_changes.append('{}: {}: {} => {}'.format(
                    section.sheet_name,
                    combination,
                    component_score,
                    combination_score))

    def _check_conflicts(self, positions, section):
        for pos, score_map in positions.items():
            for item1, item2 in itertools.combinations(
                    sorted(score_map.items()), 2):
                score1, mutation_set1 = item1
                score2, mutation_set2 = item2
                mutations1 = getattr(mutation_set1, 'mutations', None)
                mutations2 = getattr(mutation_set2, 'mutations', None)
                has_conflict = (
                        mutations1 is None or
                        mutations2 is None or
                        (mutation_set1.mutations & mutation_set2.mutations))
                if has_conflict:
                    phenotype1 = self.score_phenotypes[score1]
                    phenotype2 = self.score_phenotypes[score2]
                    self.phenotype_conflicts.append(
                        '{} in {}: {} {} => {} {}'.format(section.drug_name,
                                                          section.sheet_name,
                                                          mutation_set1,
                                                          phenotype1,
                                                          mutation_set2,
                                                          phenotype2))

    def _find_drug_summary(self, drug_summaries, drug_codes, section):
        drug_name = section.drug_name
        try:
            drug_summary = drug_summaries[drug_name]
        except KeyError:
            drug_code = drug_codes.get(drug_name)
            if drug_code is None:
                self.unknown_drugs.append('{}: {}'.format(
                    section.sheet_name,
                    drug_name))
                raise
            drug_summary = drug_summaries[drug_name] = dict(name=drug_name,
                                                            code=drug_code,
                                                            genotypes=[])
        return drug_summary

    def _monitor_positions(self, section, position_scores, reference):
        monitored_positions = getattr(section, 'monitored_positions', [])
        max_pos = len(reference.sequence)
        for pos in monitored_positions:
            scores = position_scores[pos]
            seen_variants = {mutation.variant
                             for mutation_set in scores.values()
                             for mutation in mutation_set}
            try:
                wild_type = reference.sequence[pos - 1]
            except IndexError:
                self.invalid_positions.append(
                    f'{section.sheet_name}: {section.drug_name} {pos} '
                    f'(max {max_pos})')
                continue
            mutation_set = MutationSet(
                '{}{}!{}{}'.format(wild_type,
                                   pos,
                                   ''.join(seen_variants),
                                   wild_type))
            position_scores[pos]['Effect unknown'] = mutation_set

    def _write_error_group(self, header, errors):
        if not errors:
            return
        if len(errors) == 1:
            self.errors_file.write('{}: {}.\n'.format(header, errors[0]))
        else:
            self.errors_file.write('{}s:\n  {}\n'.format(header,
                                                         '\n  '.join(errors)))

    def write_errors(self):
        missing_fold_shift_reports = []
        for (sheet_name, drug_name), mutations in sorted(
                self.missing_fold_shifts.items()):
            mutation_report = ', '.join(mutations)
            missing_fold_shift_reports.append(
                f'{sheet_name}: {drug_name} {mutation_report}')
        self._write_error_group('Unknown drug', self.unknown_drugs)
        self._write_error_group('Invalid mutation', self.invalid_mutations)
        self._write_error_group('Invalid mutation position',
                                self.invalid_mutation_positions)
        self._write_error_group('Invalid phenotype', self.invalid_phenotypes)
        self._write_error_group('Mismatched wild type', self.bad_wild_types)
        self._write_error_group('Invalid monitored position',
                                self.invalid_positions)
        self._write_error_group('Combination change', self.combination_changes)
        self._write_error_group('Missing fold shift', missing_fold_shift_reports)
        self._write_error_group('Invalid fold shift', self.invalid_fold_shifts)
        self._write_error_group('Conflicting phenotype', self.phenotype_conflicts)
        self._write_error_group('Phenotype change', self.phenotype_changes)


def main():
    args = parse_args()
    references = load_references()
    for ws in args.spreadsheet:
        ws.title = ws.title.strip()
    worksheets = [ws
                  for ws in args.spreadsheet
                  if ws.title in READY_TABS]
    fold_ranges_reader = FoldRangesReader(print_errors=False)
    reader = WorksheetReader(worksheets,
                             MonitoredPositionsReader(),
                             fold_ranges_reader)
    writer = RulesWriter(args.rules_yaml,
                         sys.stderr,
                         references,
                         fold_ranges_reader.invalid,
                         fold_ranges_reader.changes)
    writer.write(reader)
    reader.write_errors(sys.stderr)
    writer.write_errors()


def load_drug_codes():
    config_path = os.path.normpath(os.path.join(__file__,
                                                '..',
                                                '..',
                                                'resistance',
                                                'genreport.yaml'))
    with open(config_path) as config_file:
        report_config = yaml.safe_load(config_file)
        drug_codes = {}
        for section in report_config:
            known_regions = section.get('known_regions', [])
            if 'NS3' not in known_regions:
                continue
            for drug_class in section['known_drugs'].values():
                for code, name in drug_class:
                    short_name = get_short_drug_name(name)
                    drug_codes[short_name] = code
    return drug_codes


def load_references():
    projects = ProjectConfig.loadDefault()
    references = {}  # {(genotype, region): Reference}
    for ref_name, sequence in projects.getAllReferences().items():
        match = re.match(r'HCV(.*?)-.*-([^-]+)$', ref_name)
        if match:
            genotype = match.group(1)
            region = match.group(2)
            if region in HCV_REGIONS:
                reference = Reference(ref_name, sequence)
                references[(genotype, region)] = reference
                if genotype == '6':
                    references[('6E', region)] = reference
    return references


def get_short_drug_name(name):
    return name.split()[0]


def get_main_genotype(subtype):
    if 'A' <= subtype[-1] <= 'Z' or 'a' <= subtype[-1] <= 'z':
        return subtype[:-1]
    return subtype


def get_unsplit_genotype(subtype):
    main_genotype = get_main_genotype(subtype)
    if main_genotype in SPLIT_GENOTYPES:
        return main_genotype
    return subtype


def format_score(score):
    try:
        if '"' in score:
            raise ValueError('Text score contained quotes: {}.'.format(score))
        return '"{}"'.format(score)
    except TypeError:
        return score


def calculate_component_score(combination, positions):
    variant_calls = VariantCalls(combination.replace('+', ' '))
    combination_positions = {}
    for mutation_set in variant_calls:
        position_scores = positions[mutation_set.pos]
        if position_scores:
            combination_positions[mutation_set.pos] = position_scores
    if not combination_positions:
        return 0
    score_formula = build_score_formula(combination_positions)
    try:
        rule = HCVR(score_formula)
    except Exception as ex:
        raise ValueError('Bad formula for {}.'.format(combination)) from ex
    component_score = rule(variant_calls)
    return component_score


def build_score_formula(positions):
    wild_type_score = positions.pop(None, {})
    score_terms = sorted((mutation_set, format_score(score))
                         for pos, pos_scores in positions.items()
                         for score, mutation_set in pos_scores.items()
                         if score)
    for score, condition in sorted(wild_type_score.items()):
        if score:
            score_terms.insert(0, (condition, format_score(score)))
    if not score_terms:
        score_terms.append(('TRUE', '0'))
    score_formula = 'SCORE FROM ( {} )'.format(', '.join(
        '{} => {}'.format(mutation_set, score)
        for mutation_set, score in score_terms))
    return score_formula


def replace_wild_types(combination, reference):
    mutations = combination.split('+')
    for i, mutation_text in enumerate(mutations):
        mutation_set = MutationSet(mutation_text)
        expected_wild_type = reference[mutation_set.pos - 1]
        if mutation_set.wildtype != expected_wild_type:
            mutation_set = MutationSet(wildtype=expected_wild_type,
                                       pos=mutation_set.pos,
                                       variants=''.join(
                                           m.variant
                                           for m in mutation_set.mutations))
            mutations[i] = str(mutation_set)
        pass
    return '+'.join(mutations)


if __name__ == '__main__':
    main()
